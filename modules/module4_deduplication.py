"""
═══════════════════════════════════════════════════════════════════════════════
MODULE 4 — DEDUPLICATION ENGINE
Enterprise CRM Governance & Analytics Platform
═══════════════════════════════════════════════════════════════════════════════

PURPOSE:
    Consume entity_match_table from Module 3 and produce clean, de-duplicated
    master tables for Accounts, Contacts, and Leads.

HOW IT WORKS:
    1. BUILD DUPLICATE CLUSTERS
       For each entity type, entity_match_table gives (record_A → record_B)
       links.  We use Union-Find (Disjoint Set Union) to group all transitively
       linked records into one cluster.

       Example:
           A → B  and  B → C   →   cluster {A, B, C}
           D → E                →   cluster {D, E}

    2. ELECT A CANONICAL (GOLDEN) RECORD per cluster
       Priority order:
           a. If any member is already a confirmed golden_record_id → use it
           b. Else → pick the member with the lexicographically smallest ID

    3. BUILD DEDUP MAP TABLE
       One row per original record, with:
           is_duplicate     YES / NO
           canonical_id     ID of the kept record
           cluster_id       cluster identifier

    4. BUILD CLEAN MASTER TABLE
       Only the canonical record from each cluster, with:
           duplicate_count   how many records this canonical absorbed
           absorbed_ids      comma-joined absorbed source IDs

    KPIs computed:
        total_records        input row count
        canonical_records    output (deduped) row count
        duplicate_count      records identified as duplicates
        duplicate_rate       duplicate_count / total_records
        records_removed      same as duplicate_count

OUTPUTS:
    data/processed/dedup_accounts.csv
    data/processed/dedup_contacts.csv
    data/processed/dedup_leads.csv
    data/processed/dedup_map_accounts.csv
    data/processed/dedup_map_contacts.csv
    data/processed/dedup_map_leads.csv
    reports/deduplication_report.xlsx
    logs/module4_deduplication.log
"""

import os
import sys
import logging
import warnings
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings("ignore")

# ── Make the project root importable ──────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config.config import PROCESSED_DIR, LOGS_DIR, REPORTS_DIR

# ── Logger ────────────────────────────────────────────────────────────────────
os.makedirs(LOGS_DIR, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(
            os.path.join(LOGS_DIR, "module4_deduplication.log"), mode="w"
        ),
    ],
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# UNION-FIND  (Disjoint Set Union)
# ─────────────────────────────────────────────────────────────────────────────

class UnionFind:
    """
    Efficiently groups records into clusters via transitive linking.

    Usage:
        uf = UnionFind()
        uf.union("A", "B")    # A and B are the same entity
        uf.union("B", "C")    # B and C share cluster → A, B, C together
        uf.find("A")          # returns root of A's cluster
        uf.clusters()         # returns {root: [all members]}
    """

    def __init__(self):
        self._parent: dict[str, str] = {}
        self._rank:   dict[str, int] = {}

    def _make(self, x: str):
        if x not in self._parent:
            self._parent[x] = x
            self._rank[x]   = 0

    def find(self, x: str) -> str:
        self._make(x)
        if self._parent[x] != x:
            self._parent[x] = self.find(self._parent[x])   # path compression
        return self._parent[x]

    def union(self, x: str, y: str):
        rx, ry = self.find(x), self.find(y)
        if rx == ry:
            return
        # Union by rank — keeps the tree flat
        if self._rank[rx] < self._rank[ry]:
            rx, ry = ry, rx
        self._parent[ry] = rx
        if self._rank[rx] == self._rank[ry]:
            self._rank[rx] += 1

    def clusters(self) -> dict[str, list[str]]:
        result: dict[str, list[str]] = {}
        for node in self._parent:
            root = self.find(node)
            result.setdefault(root, []).append(node)
        return result


# ─────────────────────────────────────────────────────────────────────────────
# BUILD CLUSTERS
# ─────────────────────────────────────────────────────────────────────────────

def build_clusters(
    match_df: pd.DataFrame,
    entity_type: str,
    all_ids: list[str],
    golden_ids: set | None = None,
) -> pd.DataFrame:
    """
    Build a deduplication cluster map for one entity type.

    Args:
        match_df    : entity_match_table filtered to this entity_type
        entity_type : "ACCOUNT" | "CONTACT" | "LEAD"
        all_ids     : complete list of IDs from the master table
        golden_ids  : set of IDs already confirmed as canonical

    Returns:
        DataFrame with columns:
            record_id | entity_type | cluster_id | canonical_id |
            is_duplicate | cluster_size | match_type_used
    """
    uf = UnionFind()

    # Seed all known IDs so that singletons (no matches) appear in clusters
    for rid in all_ids:
        uf._make(str(rid))

    # Union all matched pairs
    for _, row in match_df.iterrows():
        uf.union(str(row["source_record_id"]), str(row["matched_record_id"]))

    clusters   = uf.clusters()
    golden_ids = golden_ids or set()
    rows: list[dict] = []

    for cluster_root, members in clusters.items():
        # Elect canonical: prefer confirmed golden ID, else lexicographic min
        golden_in_cluster = [m for m in members if m in golden_ids]
        canonical = sorted(golden_in_cluster)[0] if golden_in_cluster else sorted(members)[0]

        cluster_id = f"CLU_{entity_type[:3]}_{canonical}"

        for member in members:
            # Which match types linked this member into the cluster?
            member_links = match_df[
                (match_df["source_record_id"] == member)
                | (match_df["matched_record_id"] == member)
            ]
            match_types_used = (
                ",".join(sorted(member_links["match_type"].unique()))
                if len(member_links) > 0 else "NONE"
            )
            rows.append({
                "record_id":       member,
                "entity_type":     entity_type,
                "cluster_id":      cluster_id,
                "canonical_id":    canonical,
                "is_duplicate":    "YES" if member != canonical else "NO",
                "cluster_size":    len(members),
                "match_type_used": match_types_used,
            })

    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# BUILD DEDUP MASTER  (canonical records only)
# ─────────────────────────────────────────────────────────────────────────────

def build_dedup_master(
    master_df: pd.DataFrame,
    id_col: str,
    cluster_map: pd.DataFrame,
) -> pd.DataFrame:
    """
    Keep only canonical records; attach duplicate_count and absorbed_ids.

    Args:
        master_df   : e.g. std_accounts
        id_col      : primary-key column name
        cluster_map : output of build_clusters()

    Returns:
        dedup_df with one row per canonical entity
    """
    cm_indexed = cluster_map.set_index("record_id")

    # Count duplicates absorbed per canonical
    dupe_counts = (
        cluster_map[cluster_map["is_duplicate"] == "YES"]
        .groupby("canonical_id")
        .size()
        .rename("duplicate_count")
    )

    # Collect absorbed IDs per canonical
    absorbed = (
        cluster_map[cluster_map["is_duplicate"] == "YES"]
        .groupby("canonical_id")["record_id"]
        .apply(lambda ids: ",".join(sorted(ids)))
        .rename("absorbed_ids")
    )

    # Filter master to canonical records only
    canonical_ids = set(cluster_map[cluster_map["is_duplicate"] == "NO"]["record_id"])
    dedup_df      = master_df[master_df[id_col].isin(canonical_ids)].copy()

    dedup_df = dedup_df.merge(
        dupe_counts.reset_index().rename(columns={"canonical_id": id_col}),
        on=id_col, how="left",
    )
    dedup_df = dedup_df.merge(
        absorbed.reset_index().rename(columns={"canonical_id": id_col}),
        on=id_col, how="left",
    )
    dedup_df["duplicate_count"] = dedup_df["duplicate_count"].fillna(0).astype(int)
    dedup_df["absorbed_ids"]    = dedup_df["absorbed_ids"].fillna("")

    return dedup_df


# ─────────────────────────────────────────────────────────────────────────────
# KPI COMPUTATION
# ─────────────────────────────────────────────────────────────────────────────

def compute_kpis(
    master_df: pd.DataFrame,
    dedup_df: pd.DataFrame,
    entity_type: str,
) -> dict:
    total          = len(master_df)
    after          = len(dedup_df)
    removed        = total - after
    duplicate_rate = round(removed / total, 4) if total > 0 else 0.0
    return {
        "entity_type":      entity_type,
        "total_records":    total,
        "canonical_records": after,
        "duplicate_count":  removed,
        "duplicate_rate":   f"{duplicate_rate:.2%}",
        "records_removed":  removed,
    }


# ─────────────────────────────────────────────────────────────────────────────
# REPORT
# ─────────────────────────────────────────────────────────────────────────────

def write_deduplication_report(kpi_list: list[dict], cluster_maps: dict) -> str:
    path = os.path.join(REPORTS_DIR, "deduplication_report.xlsx")

    HDR_FILL  = PatternFill("solid", start_color="1F3864")
    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    BODY_FONT = Font(name="Arial", size=10)
    WARN_FILL = PatternFill("solid", start_color="FFC7CE")
    WARN_FONT = Font(bold=True, color="9C0006", name="Arial", size=10)
    OK_FILL   = PatternFill("solid", start_color="C6EFCE")
    OK_FONT   = Font(bold=True, color="276221", name="Arial", size=10)
    CENTER    = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="CCCCCC")
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    # ── Sheet 1: KPI Summary ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "KPI Summary"
    ws1.freeze_panes = "A2"

    h1 = ["Entity Type", "Total Records", "Canonical Records",
          "Duplicates Found", "Duplicate Rate", "Records Removed"]
    ws1.append(h1)
    for cell in ws1[1]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for kpi in kpi_list:
        ws1.append([
            kpi["entity_type"],
            kpi["total_records"],
            kpi["canonical_records"],
            kpi["duplicate_count"],
            kpi["duplicate_rate"],
            kpi["records_removed"],
        ])
        r = ws1.max_row
        dup_cell = ws1.cell(row=r, column=4)
        if kpi["duplicate_count"] > 0:
            dup_cell.fill = WARN_FILL
            dup_cell.font = WARN_FONT
        else:
            dup_cell.fill = OK_FILL
            dup_cell.font = OK_FONT
        for c in ws1[r]:
            if c.column != 4:
                c.font = BODY_FONT
            c.alignment = CENTER
            c.border = BORDER

    # Totals row
    total_records    = sum(k["total_records"]    for k in kpi_list)
    total_canonical  = sum(k["canonical_records"] for k in kpi_list)
    total_duplicates = sum(k["duplicate_count"]  for k in kpi_list)
    overall_rate     = f"{total_duplicates/total_records:.2%}" if total_records else "0.00%"
    ws1.append(["TOTAL", total_records, total_canonical,
                total_duplicates, overall_rate, total_duplicates])
    r = ws1.max_row
    for c in ws1[r]:
        c.font = Font(bold=True, name="Arial", size=10)
        c.alignment = CENTER
        c.border = BORDER

    for col, w in zip(["A","B","C","D","E","F"], [14, 14, 16, 16, 14, 14]):
        ws1.column_dimensions[col].width = w

    # ── Sheets 2-4: Cluster maps per entity ───────────────────────────────
    for entity_label, cm_df in cluster_maps.items():
        ws = wb.create_sheet(f"Clusters {entity_label}")
        ws.freeze_panes = "A2"

        h = ["record_id", "cluster_id", "canonical_id",
             "is_duplicate", "cluster_size", "match_type_used"]
        ws.append(h)
        for cell in ws[1]:
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = CENTER
            cell.border = BORDER

        for _, row in cm_df.head(3000).iterrows():
            ws.append([row[c] for c in h])
            r = ws.max_row
            dup_cell = ws.cell(row=r, column=4)
            if dup_cell.value == "YES":
                dup_cell.fill = WARN_FILL
                dup_cell.font = WARN_FONT
            else:
                dup_cell.fill = OK_FILL
                dup_cell.font = OK_FONT
            for c in ws[r]:
                if c.column != 4:
                    c.font = BODY_FONT
                c.alignment = CENTER
                c.border = BORDER

        for col, w in zip(["A","B","C","D","E","F"], [14, 22, 14, 12, 12, 18]):
            ws.column_dimensions[col].width = w

    # ── Run info ──────────────────────────────────────────────────────────
    ws_info = wb.create_sheet("Run Info")
    ws_info["A1"] = "Generated"
    ws_info["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_info["A2"] = "Module"
    ws_info["B2"] = "Module 4 — Deduplication Engine"
    ws_info["A3"] = "Platform"
    ws_info["B3"] = "Enterprise CRM Governance & Analytics Platform"
    ws_info["A4"] = "Overall Duplicate Rate"
    ws_info["B4"] = overall_rate
    for row in ws_info.iter_rows(min_row=1, max_row=4, min_col=1, max_col=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)

    wb.save(path)
    log.info(f"  Deduplication report → {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ORCHESTRATOR
# ─────────────────────────────────────────────────────────────────────────────

def run_deduplication(
    std_data: dict,
    entity_match_table: pd.DataFrame,
) -> dict:
    """
    Args:
        std_data           : dict from module2_standardization.run_standardization()
        entity_match_table : DataFrame from module3_entity_resolution.run_entity_resolution()

    Returns:
        dict {
            "dedup_accounts" : DataFrame,
            "dedup_contacts" : DataFrame,
            "dedup_leads"    : DataFrame,
        }
    """
    log.info("=" * 70)
    log.info("MODULE 4 — DEDUPLICATION ENGINE")
    log.info("=" * 70)

    os.makedirs(PROCESSED_DIR, exist_ok=True)

    std_accounts = std_data.get("std_accounts", pd.DataFrame())
    std_contacts = std_data.get("std_contacts", pd.DataFrame())
    std_leads    = std_data.get("std_leads",    pd.DataFrame())
    std_golden   = std_data.get("std_golden_map", pd.DataFrame())

    # Collect confirmed golden (canonical) account IDs from the golden map
    golden_account_ids: set[str] = set()
    if std_golden is not None and len(std_golden) > 0 and "golden_record_id" in std_golden.columns:
        for _, grp in std_golden.groupby("golden_record_id"):
            golden_account_ids.add(sorted(grp["account_id"].tolist())[0])

    kpi_list     = []
    cluster_maps = {}
    result       = {}

    # ── ACCOUNTS ──────────────────────────────────────────────────────────
    log.info("  ─── Deduplicating: ACCOUNTS")
    if len(std_accounts) > 0:
        acc_matches = entity_match_table[
            entity_match_table["entity_type"] == "ACCOUNT"
        ].copy()
        acc_cluster_map = build_clusters(
            acc_matches,
            entity_type="ACCOUNT",
            all_ids=std_accounts["account_id"].tolist(),
            golden_ids=golden_account_ids,
        )
        dedup_accounts = build_dedup_master(std_accounts, "account_id", acc_cluster_map)
        kpi_list.append(compute_kpis(std_accounts, dedup_accounts, "ACCOUNT"))
        cluster_maps["Accounts"] = acc_cluster_map

        log.info(f"    Duplicates : {kpi_list[-1]['duplicate_count']:,}"
                 f"  |  Rate : {kpi_list[-1]['duplicate_rate']}"
                 f"  |  Canonical : {len(dedup_accounts):,}")

        dedup_accounts.to_csv(
            os.path.join(PROCESSED_DIR, "dedup_accounts.csv"), index=False
        )
        acc_cluster_map.to_csv(
            os.path.join(PROCESSED_DIR, "dedup_map_accounts.csv"), index=False
        )
        result["dedup_accounts"] = dedup_accounts
    else:
        log.warning("  ⚠ std_accounts is empty — skipping account deduplication")

    # ── CONTACTS ──────────────────────────────────────────────────────────
    log.info("  ─── Deduplicating: CONTACTS")
    if len(std_contacts) > 0:
        con_matches = entity_match_table[
            entity_match_table["entity_type"] == "CONTACT"
        ].copy()
        con_cluster_map = build_clusters(
            con_matches,
            entity_type="CONTACT",
            all_ids=std_contacts["contact_id"].tolist(),
        )
        dedup_contacts = build_dedup_master(std_contacts, "contact_id", con_cluster_map)
        kpi_list.append(compute_kpis(std_contacts, dedup_contacts, "CONTACT"))
        cluster_maps["Contacts"] = con_cluster_map

        log.info(f"    Duplicates : {kpi_list[-1]['duplicate_count']:,}"
                 f"  |  Rate : {kpi_list[-1]['duplicate_rate']}"
                 f"  |  Canonical : {len(dedup_contacts):,}")

        dedup_contacts.to_csv(
            os.path.join(PROCESSED_DIR, "dedup_contacts.csv"), index=False
        )
        con_cluster_map.to_csv(
            os.path.join(PROCESSED_DIR, "dedup_map_contacts.csv"), index=False
        )
        result["dedup_contacts"] = dedup_contacts
    else:
        log.warning("  ⚠ std_contacts is empty — skipping contact deduplication")

    # ── LEADS ─────────────────────────────────────────────────────────────
    log.info("  ─── Deduplicating: LEADS")
    if len(std_leads) > 0:
        lead_matches = entity_match_table[
            entity_match_table["entity_type"] == "LEAD"
        ].copy()
        lead_cluster_map = build_clusters(
            lead_matches,
            entity_type="LEAD",
            all_ids=std_leads["lead_id"].tolist(),
        )
        dedup_leads = build_dedup_master(std_leads, "lead_id", lead_cluster_map)
        kpi_list.append(compute_kpis(std_leads, dedup_leads, "LEAD"))
        cluster_maps["Leads"] = lead_cluster_map

        log.info(f"    Duplicates : {kpi_list[-1]['duplicate_count']:,}"
                 f"  |  Rate : {kpi_list[-1]['duplicate_rate']}"
                 f"  |  Canonical : {len(dedup_leads):,}")

        dedup_leads.to_csv(
            os.path.join(PROCESSED_DIR, "dedup_leads.csv"), index=False
        )
        lead_cluster_map.to_csv(
            os.path.join(PROCESSED_DIR, "dedup_map_leads.csv"), index=False
        )
        result["dedup_leads"] = dedup_leads
    else:
        log.warning("  ⚠ std_leads is empty — skipping lead deduplication")

    # ── Report + Summary ──────────────────────────────────────────────────
    log.info("-" * 70)
    log.info("KPI SUMMARY:")
    total_removed = 0
    for kpi in kpi_list:
        log.info(
            f"  {kpi['entity_type']:<10} | Total: {kpi['total_records']:>8,}"
            f"  | Dupes: {kpi['duplicate_count']:>6,}"
            f"  | Rate: {kpi['duplicate_rate']}"
            f"  | Canonical: {kpi['canonical_records']:>7,}"
        )
        total_removed += kpi["records_removed"]
    log.info(f"  TOTAL records removed : {total_removed:,}")
    log.info("-" * 70)

    if kpi_list:
        write_deduplication_report(kpi_list, cluster_maps)

    log.info("=" * 70)
    log.info("MODULE 4 COMPLETE")
    log.info("=" * 70)

    return result


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    from module1_ingestion import run_ingestion
    from module2_standardization import run_standardization
    from module3_entity_resolution import run_entity_resolution

    staging = run_ingestion()
    std     = run_standardization(staging)
    emt     = run_entity_resolution(std)
    run_deduplication(std, emt)
