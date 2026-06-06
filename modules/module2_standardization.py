"""
═══════════════════════════════════════════════════════════════════════════════
MODULE 2 — DATA STANDARDIZATION LAYER
Enterprise CRM Governance & Analytics Platform
═══════════════════════════════════════════════════════════════════════════════

PURPOSE:
    Take every staging DataFrame produced by Module 1 and convert all
    inconsistent values into clean, business-standard formats.

TRANSFORMATIONS:
    Accounts  — industry mapping, country mapping, revenue / employee typing
    Contacts  — email validation, full-name creation, job-title normalisation
    Leads     — source normalisation, status normalisation
    Stewardship — DQ status normalisation
    Golden Map  — ID format validation

OUTPUTS:
    data/processed/std_accounts.csv
    data/processed/std_contacts.csv
    data/processed/std_leads.csv
    data/processed/std_stewardship.csv
    data/processed/std_golden_map.csv

    reports/standardization_report.xlsx  ← change summary per table
    logs/module2_standardization.log
"""

import os
import re
import sys
import logging
import warnings
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings("ignore")

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config.config import (
    INDUSTRY_MAP, COUNTRY_MAP,
    VALID_LEAD_SOURCES, VALID_LEAD_STATUSES,
    VALID_JOB_TITLES, VALID_DQ_STATUSES,
    PROCESSED_DIR, LOGS_DIR, REPORTS_DIR,
)

# ── Logger ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(os.path.join(LOGS_DIR, "module2_standardization.log"), mode="w"),
    ]
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# HELPER UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

EMAIL_REGEX = re.compile(
    r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$"
)


def is_valid_email(value) -> bool:
    if pd.isna(value) or not isinstance(value, str):
        return False
    return bool(EMAIL_REGEX.match(value.strip()))


def normalise_phone(value) -> str | None:
    """Strip all non-digit characters; return None if < 7 digits."""
    if pd.isna(value) or not isinstance(value, str):
        return None
    digits = re.sub(r"\D", "", value)
    return digits if len(digits) >= 7 else None


def map_lookup(value, mapping: dict, default: str = "Other") -> str:
    """Case-insensitive dict lookup with a fallback default."""
    if pd.isna(value) or not isinstance(value, str):
        return default
    key = value.strip().lower()
    return mapping.get(key, value.strip())   # keep original if not in map


def normalise_to_valid_list(value, valid_list: list[str], default: str = "Other") -> str:
    """Return value if it is in valid_list (case-insensitive), else default."""
    if pd.isna(value) or not isinstance(value, str):
        return default
    for valid in valid_list:
        if value.strip().lower() == valid.lower():
            return valid
    return default


def title_case_name(value) -> str | None:
    if pd.isna(value) or not isinstance(value, str):
        return None
    return value.strip().title()


def count_changes(original: pd.Series, updated: pd.Series) -> int:
    """Count the number of cells that changed between two Series."""
    return int((original.fillna("") != updated.fillna("")).sum())


# ─────────────────────────────────────────────────────────────────────────────
# ACCOUNTS STANDARDISATION
# ─────────────────────────────────────────────────────────────────────────────

def standardise_accounts(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """
    Rules:
      - industry  → canonical value from INDUSTRY_MAP
      - country   → canonical value from COUNTRY_MAP
      - account_name → title case, strip whitespace
      - annual_revenue / employee_count → numeric (already cast in M1, but re-cast defensively)
      - std_flag  → NEW column: 'OK' if all key fields are clean, else 'REVIEW'
    """
    log.info("  Standardising: Accounts")
    out = df.copy()
    changes = {}

    # Industry
    orig = out["industry"].copy()
    out["industry"] = out["industry"].apply(lambda v: map_lookup(v, INDUSTRY_MAP, default=v))
    changes["industry_changes"] = count_changes(orig, out["industry"])

    # Country
    orig = out["country"].copy()
    out["country"] = out["country"].apply(lambda v: map_lookup(v, COUNTRY_MAP, default=v))
    changes["country_changes"] = count_changes(orig, out["country"])

    # Account name — title case
    orig = out["account_name"].copy()
    out["account_name"] = out["account_name"].apply(title_case_name)
    changes["name_changes"] = count_changes(orig, out["account_name"])

    # Numeric columns
    out["annual_revenue"]  = pd.to_numeric(out["annual_revenue"],  errors="coerce")
    out["employee_count"]  = pd.to_numeric(out["employee_count"],  errors="coerce")

    # Flag: REVIEW if industry or country is still unknown-ish
    known_industries = set(INDUSTRY_MAP.values())
    known_countries  = set(COUNTRY_MAP.values())
    out["std_flag"] = out.apply(
        lambda r: "REVIEW"
        if (r["industry"] not in known_industries or r["country"] not in known_countries)
        else "OK",
        axis=1,
    )
    changes["review_flagged"] = int((out["std_flag"] == "REVIEW").sum())

    log.info(f"    Industry changes: {changes['industry_changes']}"
             f"  | Country changes: {changes['country_changes']}"
             f"  | Flagged for review: {changes['review_flagged']}")

    changes["table"]      = "std_accounts"
    changes["total_rows"] = len(out)
    return out, changes


# ─────────────────────────────────────────────────────────────────────────────
# CONTACTS STANDARDISATION
# ─────────────────────────────────────────────────────────────────────────────

def standardise_contacts(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """
    Rules:
      - first_name / last_name → title case
      - full_name  → NEW column: first + last
      - email      → validate format; add email_valid flag
      - job_title  → normalise against VALID_JOB_TITLES
    """
    log.info("  Standardising: Contacts")
    out = df.copy()
    changes = {}

    # Name casing
    orig_fn = out["first_name"].copy()
    orig_ln = out["last_name"].copy()
    out["first_name"] = out["first_name"].apply(title_case_name)
    out["last_name"]  = out["last_name"].apply(title_case_name)
    changes["name_changes"] = count_changes(orig_fn, out["first_name"]) + \
                               count_changes(orig_ln, out["last_name"])

    # Full name
    out["full_name"] = out["first_name"].fillna("") + " " + out["last_name"].fillna("")
    out["full_name"] = out["full_name"].str.strip()

    # Email validation
    out["email_valid"] = out["email"].apply(is_valid_email).map({True: "YES", False: "NO"})
    changes["invalid_emails"] = int((out["email_valid"] == "NO").sum())

    # Job title normalisation
    orig_jt = out["job_title"].copy()
    out["job_title"] = out["job_title"].apply(
        lambda v: normalise_to_valid_list(v, VALID_JOB_TITLES, default="Other")
    )
    changes["job_title_changes"] = count_changes(orig_jt, out["job_title"])

    log.info(f"    Name changes: {changes['name_changes']}"
             f"  | Invalid emails: {changes['invalid_emails']}"
             f"  | Job title changes: {changes['job_title_changes']}")

    changes["table"]      = "std_contacts"
    changes["total_rows"] = len(out)
    return out, changes


# ─────────────────────────────────────────────────────────────────────────────
# LEADS STANDARDISATION
# ─────────────────────────────────────────────────────────────────────────────

def standardise_leads(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """
    Rules:
      - source → normalise against VALID_LEAD_SOURCES
      - status → normalise against VALID_LEAD_STATUSES
      - lead_name → title case
    """
    log.info("  Standardising: Leads")
    out = df.copy()
    changes = {}

    orig_src = out["source"].copy()
    out["source"] = out["source"].apply(
        lambda v: normalise_to_valid_list(v, VALID_LEAD_SOURCES, default="Other")
    )
    changes["source_changes"] = count_changes(orig_src, out["source"])

    orig_sts = out["status"].copy()
    out["status"] = out["status"].apply(
        lambda v: normalise_to_valid_list(v, VALID_LEAD_STATUSES, default="New")
    )
    changes["status_changes"] = count_changes(orig_sts, out["status"])

    orig_nm = out["lead_name"].copy()
    out["lead_name"] = out["lead_name"].apply(title_case_name)
    changes["name_changes"] = count_changes(orig_nm, out["lead_name"])

    log.info(f"    Source changes: {changes['source_changes']}"
             f"  | Status changes: {changes['status_changes']}")

    changes["table"]      = "std_leads"
    changes["total_rows"] = len(out)
    return out, changes


# ─────────────────────────────────────────────────────────────────────────────
# STEWARDSHIP STANDARDISATION
# ─────────────────────────────────────────────────────────────────────────────

def standardise_stewardship(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """
    Rules:
      - dq_issue_status → normalise against VALID_DQ_STATUSES
      - steward → title case
    """
    log.info("  Standardising: Data_Stewardship")
    out = df.copy()
    changes = {}

    orig = out["dq_issue_status"].copy()
    out["dq_issue_status"] = out["dq_issue_status"].apply(
        lambda v: normalise_to_valid_list(v, VALID_DQ_STATUSES, default="Open")
    )
    changes["status_changes"] = count_changes(orig, out["dq_issue_status"])

    orig_s = out["steward"].copy()
    out["steward"] = out["steward"].apply(title_case_name)
    changes["steward_changes"] = count_changes(orig_s, out["steward"])

    log.info(f"    Status changes: {changes['status_changes']}"
             f"  | Steward changes: {changes['steward_changes']}")

    changes["table"]      = "std_stewardship"
    changes["total_rows"] = len(out)
    return out, changes


# ─────────────────────────────────────────────────────────────────────────────
# GOLDEN MAP STANDARDISATION
# ─────────────────────────────────────────────────────────────────────────────

def standardise_golden_map(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """
    Rules:
      - source_record, account_id, golden_record_id → strip whitespace,
        uppercase to ensure ID consistency
      - id_format_valid → flag IDs that don't follow expected patterns
    """
    log.info("  Standardising: Golden_Record_Map")
    out = df.copy()
    changes = {}

    for col in ["source_record", "account_id", "golden_record_id"]:
        out[col] = out[col].str.strip().str.upper()

    # Validate ID patterns: ACC######, DUP#####, GOLD#####
    acc_pattern   = re.compile(r"^ACC\d{6}$")
    dup_pattern   = re.compile(r"^DUP\d{5}$")
    gold_pattern  = re.compile(r"^GOLD\d{5}$")

    def check_ids(row):
        ok = (
            bool(acc_pattern.match(str(row["account_id"])))
            and bool(dup_pattern.match(str(row["source_record"])))
            and bool(gold_pattern.match(str(row["golden_record_id"])))
        )
        return "VALID" if ok else "INVALID"

    out["id_format_valid"] = out.apply(check_ids, axis=1)
    changes["invalid_id_formats"] = int((out["id_format_valid"] == "INVALID").sum())

    log.info(f"    Invalid ID formats: {changes['invalid_id_formats']}")

    changes["table"]      = "std_golden_map"
    changes["total_rows"] = len(out)
    return out, changes


# ─────────────────────────────────────────────────────────────────────────────
# WRITE PROCESSED FILES
# ─────────────────────────────────────────────────────────────────────────────

def write_processed(name: str, df: pd.DataFrame) -> str:
    os.makedirs(PROCESSED_DIR, exist_ok=True)
    path = os.path.join(PROCESSED_DIR, f"{name}.csv")
    df.to_csv(path, index=False)
    log.info(f"    Saved → {path}  ({len(df):,} rows)")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# STANDARDIZATION REPORT
# ─────────────────────────────────────────────────────────────────────────────

def write_standardization_report(all_changes: list[dict]) -> str:
    """
    Excel report with two sheets:
      Summary   — one row per table, key change counts
      Details   — expandable KPI breakdown
    """
    path = os.path.join(REPORTS_DIR, "standardization_report.xlsx")

    HDR_FILL = PatternFill("solid", start_color="1F3864")
    HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    BODY_FONT = Font(name="Arial", size=10)
    WARN_FILL = PatternFill("solid", start_color="FFEB9C")
    WARN_FONT = Font(bold=True, color="9C5700", name="Arial", size=10)
    OK_FILL   = PatternFill("solid", start_color="C6EFCE")
    OK_FONT   = Font(bold=True, color="276221", name="Arial", size=10)
    CENTER    = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="CCCCCC")
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.freeze_panes = "A2"

    headers = ["Table", "Total Rows", "Changes Made", "Notes"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for c in all_changes:
        table      = c.get("table", "—")
        total_rows = c.get("total_rows", 0)
        # Sum all numeric change counts (exclude table & total_rows keys)
        change_sum = sum(v for k, v in c.items()
                         if k not in ("table", "total_rows") and isinstance(v, int))
        notes = "; ".join(
            f"{k.replace('_', ' ').title()}: {v}"
            for k, v in c.items()
            if k not in ("table", "total_rows") and isinstance(v, int)
        )
        ws.append([table, total_rows, change_sum, notes])
        r = ws.max_row
        change_cell = ws.cell(row=r, column=3)
        if change_sum > 0:
            change_cell.fill = WARN_FILL
            change_cell.font = WARN_FONT
        else:
            change_cell.fill = OK_FILL
            change_cell.font = OK_FONT
        for cell in ws[r]:
            if cell.column != 3:
                cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 80

    # ── Meta sheet ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Run Info")
    ws2["A1"] = "Report Generated"
    ws2["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws2["A2"] = "Module"
    ws2["B2"] = "Module 2 — Data Standardization"
    ws2["A3"] = "Platform"
    ws2["B3"] = "Enterprise CRM Governance & Analytics Platform"
    for row in ws2.iter_rows(min_row=1, max_row=3, min_col=1, max_col=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)

    wb.save(path)
    log.info(f"  Standardization report saved → {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ORCHESTRATOR
# ─────────────────────────────────────────────────────────────────────────────

def run_standardization(staging_data: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    """
    End-to-end Module 2 execution.

    Args:
        staging_data:  dict returned by module1_ingestion.run_ingestion()
                       keys = staging table names (e.g. 'stg_accounts')

    Returns:
        dict {std_table_name: DataFrame}
    """
    log.info("=" * 70)
    log.info("MODULE 2 — DATA STANDARDIZATION LAYER")
    log.info("=" * 70)
    os.makedirs(PROCESSED_DIR, exist_ok=True)

    std_data   = {}
    all_changes = []

    # Map staging names → standardisation functions → output names
    pipeline = [
        ("stg_accounts",    standardise_accounts,     "std_accounts"),
        ("stg_contacts",    standardise_contacts,      "std_contacts"),
        ("stg_leads",       standardise_leads,         "std_leads"),
        ("stg_stewardship", standardise_stewardship,   "std_stewardship"),
        ("stg_golden_map",  standardise_golden_map,    "std_golden_map"),
    ]

    for staging_name, fn, output_name in pipeline:
        if staging_name not in staging_data:
            log.warning(f"  ⚠ {staging_name} not found in staging data — skipping")
            continue

        log.info(f"  ─── {staging_name}  →  {output_name}")
        df_std, changes = fn(staging_data[staging_name])
        write_processed(output_name, df_std)
        std_data[output_name] = df_std
        all_changes.append(changes)

    # Deliverable
    log.info("-" * 70)
    log.info("Writing deliverables...")
    write_standardization_report(all_changes)

    log.info("=" * 70)
    log.info(f"MODULE 2 COMPLETE — {len(std_data)} tables standardised")
    log.info("=" * 70)

    return std_data


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # Run Module 2 standalone: needs Module 1 output in data/staging/
    from module1_ingestion import run_ingestion
    staging = run_ingestion()
    run_standardization(staging)
