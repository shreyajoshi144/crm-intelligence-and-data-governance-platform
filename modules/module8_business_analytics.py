"""
MODULE 8 - BUSINESS ANALYTICS LAYER
Enterprise CRM Governance & Analytics Platform

Purpose
-------
Create business-facing analytics datasets using trusted Golden Records.

Audience:
    - Sales Leadership
    - Business Development Teams
    - Marketing Teams
    - Executive Leadership

This module focuses on business performance metrics, not governance metrics.

Outputs:
    data/processed/business_analytics_dataset.csv
    data/processed/lead_funnel_summary.csv
    data/processed/lead_source_performance.csv
    data/processed/industry_distribution.csv
    data/processed/country_distribution.csv
    reports/business_analytics_report.xlsx
    logs/module8_business_analytics.log

Public entry point:
    run_business_analytics(golden_data=None) -> dict
"""

from __future__ import annotations

import logging
import os
import sys
import warnings
from datetime import date, datetime
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

warnings.filterwarnings("ignore")

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from config.config import LOGS_DIR, PROCESSED_DIR, REPORTS_DIR
except Exception:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    PROCESSED_DIR = os.path.join(BASE_DIR, "data", "processed")
    LOGS_DIR = os.path.join(BASE_DIR, "logs")
    REPORTS_DIR = os.path.join(BASE_DIR, "reports")


for directory in (PROCESSED_DIR, LOGS_DIR, REPORTS_DIR):
    os.makedirs(directory, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(os.path.join(LOGS_DIR, "module8_business_analytics.log"), mode="w"),
    ],
)
log = logging.getLogger(__name__)


REPORTING_DATE = date.today().isoformat()
LEAD_STATUSES = ["New", "Working", "Qualified", "Converted", "Unqualified", "Nurturing"]
LEAD_SOURCES = ["LinkedIn", "Referral", "Event", "Web", "Partner", "Advertisement", "Cold Call", "Other"]


def _csv_path(filename: str) -> str:
    return os.path.join(PROCESSED_DIR, filename)


def _read_processed_csv(filename: str) -> pd.DataFrame:
    path = _csv_path(filename)
    if not os.path.exists(path):
        log.warning("  Missing optional input file: %s", path)
        return pd.DataFrame()
    df = pd.read_csv(path)
    log.info("  Loaded %-34s %8s rows", filename, f"{len(df):,}")
    return df


def _frame_from_source(source: dict[str, Any] | None, key: str, filename: str) -> pd.DataFrame:
    if isinstance(source, dict) and key in source and source[key] is not None:
        return source[key].copy()
    return _read_processed_csv(filename)


def _safe_str(value: Any, default: str = "Unknown") -> str:
    if pd.isna(value):
        return default
    text = str(value).strip()
    return text if text else default


def _safe_numeric_series(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def _percentage(count: int | float, total: int | float) -> float:
    if not total:
        return 0.0
    return round((count / total) * 100, 2)


def _metric(
    category: str,
    name: str,
    value: Any,
    dimension: str = "Overall",
    dimension_value: str = "Overall",
) -> dict[str, Any]:
    return {
        "metric_category": category,
        "metric_name": name,
        "metric_value": value,
        "dimension": dimension,
        "dimension_value": dimension_value,
        "reporting_date": REPORTING_DATE,
    }


def build_lead_funnel(leads: pd.DataFrame) -> pd.DataFrame:
    total = len(leads)
    if len(leads) == 0 or "status" not in leads.columns:
        return pd.DataFrame({
            "status": LEAD_STATUSES,
            "lead_count": [0] * len(LEAD_STATUSES),
            "percentage": [0.0] * len(LEAD_STATUSES),
            "funnel_order": list(range(1, len(LEAD_STATUSES) + 1)),
        })

    counts = leads["status"].fillna("Unknown").astype(str).str.strip().value_counts()
    rows = []
    for order, status in enumerate(LEAD_STATUSES, start=1):
        count = int(counts.get(status, 0))
        rows.append({
            "status": status,
            "lead_count": count,
            "percentage": _percentage(count, total),
            "funnel_order": order,
        })

    extra_statuses = sorted(set(counts.index) - set(LEAD_STATUSES))
    for status in extra_statuses:
        count = int(counts.get(status, 0))
        rows.append({
            "status": status,
            "lead_count": count,
            "percentage": _percentage(count, total),
            "funnel_order": len(rows) + 1,
        })

    return pd.DataFrame(rows)


def build_lead_source_performance(leads: pd.DataFrame) -> pd.DataFrame:
    columns = ["source", "lead_count", "qualified_leads", "converted_leads", "conversion_rate"]
    if len(leads) == 0:
        return pd.DataFrame(columns=columns)

    working = leads.copy()
    if "source" not in working.columns:
        working["source"] = "Unknown"
    if "status" not in working.columns:
        working["status"] = "Unknown"

    working["source"] = working["source"].apply(_safe_str)
    working["status"] = working["status"].apply(_safe_str)

    rows = []
    source_order = LEAD_SOURCES + sorted(set(working["source"]) - set(LEAD_SOURCES))
    for source in source_order:
        group = working[working["source"].eq(source)]
        if len(group) == 0:
            continue
        lead_count = len(group)
        qualified = int(group["status"].eq("Qualified").sum())
        converted = int(group["status"].eq("Converted").sum())
        rows.append({
            "source": source,
            "lead_count": lead_count,
            "qualified_leads": qualified,
            "converted_leads": converted,
            "conversion_rate": _percentage(converted, lead_count),
        })

    return pd.DataFrame(rows, columns=columns)


def build_distribution(df: pd.DataFrame, column: str, label_col: str, count_col: str) -> pd.DataFrame:
    if len(df) == 0 or column not in df.columns:
        return pd.DataFrame(columns=[label_col, count_col, "percentage"])

    total = len(df)
    counts = df[column].apply(_safe_str).value_counts().reset_index()
    counts.columns = [label_col, count_col]
    counts["percentage"] = counts[count_col].apply(lambda count: _percentage(count, total))
    return counts.sort_values([count_col, label_col], ascending=[False, True]).reset_index(drop=True)


def build_account_size_distribution(accounts: pd.DataFrame) -> pd.DataFrame:
    columns = ["account_size", "account_count", "percentage"]
    if len(accounts) == 0 or "employee_count" not in accounts.columns:
        return pd.DataFrame(columns=columns)

    employees = _safe_numeric_series(accounts["employee_count"])

    def bucket(value: float) -> str:
        if pd.isna(value) or value <= 0:
            return "Unknown"
        if value <= 100:
            return "Small"
        if value <= 1000:
            return "Medium"
        return "Large"

    bucketed = employees.apply(bucket)
    total = len(accounts)
    order = ["Small", "Medium", "Large", "Unknown"]
    rows = []
    for label in order:
        count = int(bucketed.eq(label).sum())
        if count == 0 and label == "Unknown":
            continue
        rows.append({
            "account_size": label,
            "account_count": count,
            "percentage": _percentage(count, total),
        })
    return pd.DataFrame(rows, columns=columns)


def build_revenue_distribution(accounts: pd.DataFrame) -> pd.DataFrame:
    columns = ["revenue_bucket", "account_count", "percentage"]
    if len(accounts) == 0 or "annual_revenue" not in accounts.columns:
        return pd.DataFrame(columns=columns)

    revenue = _safe_numeric_series(accounts["annual_revenue"])

    def bucket(value: float) -> str:
        if pd.isna(value) or value < 0:
            return "Unknown"
        if value < 1_000_000:
            return "< 1M"
        if value < 10_000_000:
            return "1M - 10M"
        if value < 100_000_000:
            return "10M - 100M"
        return "100M+"

    bucketed = revenue.apply(bucket)
    total = len(accounts)
    order = ["< 1M", "1M - 10M", "10M - 100M", "100M+", "Unknown"]
    rows = []
    for label in order:
        count = int(bucketed.eq(label).sum())
        if count == 0 and label == "Unknown":
            continue
        rows.append({
            "revenue_bucket": label,
            "account_count": count,
            "percentage": _percentage(count, total),
        })
    return pd.DataFrame(rows, columns=columns)


def _build_account_industry_lookup(accounts: pd.DataFrame) -> dict[str, str]:
    lookup: dict[str, str] = {}
    if len(accounts) == 0 or "industry" not in accounts.columns:
        return lookup

    id_columns = [
        col for col in [
            "golden_account_id", "surviving_account_id", "account_id",
        ] if col in accounts.columns
    ]

    for _, row in accounts.iterrows():
        industry = _safe_str(row.get("industry"))
        for col in id_columns:
            record_id = _safe_str(row.get(col), "")
            if record_id:
                lookup[record_id] = industry
        if "source_account_ids" in accounts.columns:
            for record_id in str(row.get("source_account_ids", "")).split(","):
                record_id = record_id.strip()
                if record_id:
                    lookup[record_id] = industry
    return lookup


def build_top_industries_by_leads(leads: pd.DataFrame, accounts: pd.DataFrame) -> pd.DataFrame:
    columns = ["industry", "lead_count", "percentage"]
    if len(leads) == 0:
        return pd.DataFrame(columns=columns)

    working = leads.copy()
    if "industry" not in working.columns:
        account_lookup = _build_account_industry_lookup(accounts)
        join_col = next((col for col in ["account_id", "surviving_account_id", "golden_account_id"] if col in working.columns), None)
        if join_col:
            working["industry"] = working[join_col].astype(str).str.strip().map(account_lookup).fillna("Unknown")
        else:
            working["industry"] = "Unknown"

    return build_distribution(working, "industry", "industry", "lead_count")


def build_executive_summary(
    accounts: pd.DataFrame,
    contacts: pd.DataFrame,
    leads: pd.DataFrame,
    source_perf: pd.DataFrame,
    industry_distribution: pd.DataFrame,
    country_distribution: pd.DataFrame,
) -> pd.DataFrame:
    total_accounts = len(accounts)
    total_contacts = len(contacts)
    total_leads = len(leads)
    converted_leads = int(leads["status"].eq("Converted").sum()) if "status" in leads.columns else 0
    conversion_rate = _percentage(converted_leads, total_leads)

    top_source = (
        source_perf.sort_values(["converted_leads", "lead_count", "source"], ascending=[False, False, True]).iloc[0]["source"]
        if len(source_perf) > 0 else "Unknown"
    )
    top_industry = industry_distribution.iloc[0]["industry"] if len(industry_distribution) > 0 else "Unknown"
    top_country = country_distribution.iloc[0]["country"] if len(country_distribution) > 0 else "Unknown"

    rows = [
        ("Total Accounts", total_accounts),
        ("Total Contacts", total_contacts),
        ("Total Leads", total_leads),
        ("Converted Leads", converted_leads),
        ("Conversion Rate", conversion_rate),
        ("Top Lead Source", top_source),
        ("Top Industry", top_industry),
        ("Top Country", top_country),
    ]
    return pd.DataFrame(rows, columns=["metric_name", "metric_value"])


def build_business_analytics_dataset(
    executive_summary: pd.DataFrame,
    lead_funnel: pd.DataFrame,
    source_perf: pd.DataFrame,
    industry_distribution: pd.DataFrame,
    country_distribution: pd.DataFrame,
    size_distribution: pd.DataFrame,
    revenue_distribution: pd.DataFrame,
    job_title_distribution: pd.DataFrame,
    leads_by_industry: pd.DataFrame,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []

    for _, row in executive_summary.iterrows():
        rows.append(_metric("Executive Summary", row["metric_name"], row["metric_value"]))

    for _, row in lead_funnel.iterrows():
        rows.append(_metric("Lead Funnel", "Lead Count", row["lead_count"], "Status", row["status"]))
        rows.append(_metric("Lead Funnel", "Percentage", row["percentage"], "Status", row["status"]))

    for _, row in source_perf.iterrows():
        rows.append(_metric("Lead Source", "Lead Count", row["lead_count"], "Source", row["source"]))
        rows.append(_metric("Lead Source", "Qualified Leads", row["qualified_leads"], "Source", row["source"]))
        rows.append(_metric("Lead Source", "Converted Leads", row["converted_leads"], "Source", row["source"]))
        rows.append(_metric("Lead Source", "Conversion Rate", row["conversion_rate"], "Source", row["source"]))

    for _, row in industry_distribution.iterrows():
        rows.append(_metric("Industry Distribution", "Account Count", row["account_count"], "Industry", row["industry"]))
        rows.append(_metric("Industry Distribution", "Percentage", row["percentage"], "Industry", row["industry"]))

    for _, row in country_distribution.iterrows():
        rows.append(_metric("Country Distribution", "Account Count", row["account_count"], "Country", row["country"]))
        rows.append(_metric("Country Distribution", "Percentage", row["percentage"], "Country", row["country"]))

    for _, row in size_distribution.iterrows():
        rows.append(_metric("Account Size", "Account Count", row["account_count"], "Account Size", row["account_size"]))
        rows.append(_metric("Account Size", "Percentage", row["percentage"], "Account Size", row["account_size"]))

    for _, row in revenue_distribution.iterrows():
        rows.append(_metric("Revenue Distribution", "Account Count", row["account_count"], "Revenue Bucket", row["revenue_bucket"]))
        rows.append(_metric("Revenue Distribution", "Percentage", row["percentage"], "Revenue Bucket", row["revenue_bucket"]))

    for _, row in job_title_distribution.iterrows():
        rows.append(_metric("Contact Job Title", "Contact Count", row["contact_count"], "Job Title", row["job_title"]))
        rows.append(_metric("Contact Job Title", "Percentage", row["percentage"], "Job Title", row["job_title"]))

    for _, row in leads_by_industry.iterrows():
        rows.append(_metric("Top Industries by Leads", "Lead Count", row["lead_count"], "Industry", row["industry"]))
        rows.append(_metric("Top Industries by Leads", "Percentage", row["percentage"], "Industry", row["industry"]))

    return pd.DataFrame(rows, columns=[
        "metric_category", "metric_name", "metric_value",
        "dimension", "dimension_value", "reporting_date",
    ])


def write_business_analytics_report(
    executive_summary: pd.DataFrame,
    lead_funnel: pd.DataFrame,
    source_perf: pd.DataFrame,
    industry_distribution: pd.DataFrame,
    country_distribution: pd.DataFrame,
    size_distribution: pd.DataFrame,
    revenue_distribution: pd.DataFrame,
    job_title_distribution: pd.DataFrame,
    accounts: pd.DataFrame,
    contacts: pd.DataFrame,
    leads: pd.DataFrame,
) -> str:
    path = os.path.join(REPORTS_DIR, "business_analytics_report.xlsx")

    header_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    body_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def write_df(ws, df: pd.DataFrame) -> None:
        ws.freeze_panes = "A2"
        ws.append(list(df.columns))
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        for _, row in df.iterrows():
            ws.append([row[col] for col in df.columns])
            for cell in ws[ws.max_row]:
                cell.font = body_font
                cell.alignment = left if isinstance(cell.value, str) and len(str(cell.value)) > 25 else center
                cell.border = border

        for column_cells in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 52)

    account_segmentation = pd.concat([
        size_distribution.assign(segment_type="Account Size").rename(columns={"account_size": "segment", "account_count": "record_count"}),
        revenue_distribution.assign(segment_type="Revenue").rename(columns={"revenue_bucket": "segment", "account_count": "record_count"}),
    ], ignore_index=True)
    if len(account_segmentation) > 0:
        account_segmentation = account_segmentation[["segment_type", "segment", "record_count", "percentage"]]
    else:
        account_segmentation = pd.DataFrame(columns=["segment_type", "segment", "record_count", "percentage"])

    contact_segmentation = job_title_distribution.rename(columns={
        "job_title": "segment",
        "contact_count": "record_count",
    })
    if len(contact_segmentation) > 0:
        contact_segmentation.insert(0, "segment_type", "Job Title")
        contact_segmentation = contact_segmentation[["segment_type", "segment", "record_count", "percentage"]]
    else:
        contact_segmentation = pd.DataFrame(columns=["segment_type", "segment", "record_count", "percentage"])

    run_information = pd.DataFrame([
        {"item": "Timestamp", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"item": "Module Name", "value": "Module 8 - Business Analytics"},
        {"item": "Reporting Date", "value": REPORTING_DATE},
        {"item": "Golden Account Records", "value": len(accounts)},
        {"item": "Golden Contact Records", "value": len(contacts)},
        {"item": "Golden Lead Records", "value": len(leads)},
    ])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    write_df(ws, executive_summary)

    ws = wb.create_sheet("Lead Funnel")
    write_df(ws, lead_funnel)

    ws = wb.create_sheet("Lead Source Performance")
    write_df(ws, source_perf)

    ws = wb.create_sheet("Industry Distribution")
    write_df(ws, industry_distribution)

    ws = wb.create_sheet("Country Distribution")
    write_df(ws, country_distribution)

    ws = wb.create_sheet("Account Segmentation")
    write_df(ws, account_segmentation)

    ws = wb.create_sheet("Contact Segmentation")
    write_df(ws, contact_segmentation)

    ws = wb.create_sheet("Run Information")
    write_df(ws, run_information)

    wb.save(path)
    log.info("  Business analytics report -> %s", path)
    return path


def run_business_analytics(golden_data: dict[str, pd.DataFrame] | None = None) -> dict[str, Any]:
    log.info("=" * 70)
    log.info("MODULE 8 - BUSINESS ANALYTICS")
    log.info("=" * 70)

    golden_data = golden_data or {}

    accounts = _frame_from_source(golden_data, "golden_accounts", "golden_accounts.csv")
    contacts = _frame_from_source(golden_data, "golden_contacts", "golden_contacts.csv")
    leads = _frame_from_source(golden_data, "golden_leads", "golden_leads.csv")

    log.info(
        "  Input: %s golden accounts | %s golden contacts | %s golden leads",
        f"{len(accounts):,}", f"{len(contacts):,}", f"{len(leads):,}",
    )

    lead_funnel = build_lead_funnel(leads)
    source_perf = build_lead_source_performance(leads)
    industry_distribution = build_distribution(accounts, "industry", "industry", "account_count")
    country_distribution = build_distribution(accounts, "country", "country", "account_count")
    size_distribution = build_account_size_distribution(accounts)
    revenue_distribution = build_revenue_distribution(accounts)
    job_title_distribution = build_distribution(contacts, "job_title", "job_title", "contact_count")
    leads_by_industry = build_top_industries_by_leads(leads, accounts)

    executive_summary = build_executive_summary(
        accounts,
        contacts,
        leads,
        source_perf,
        industry_distribution,
        country_distribution,
    )

    analytics_df = build_business_analytics_dataset(
        executive_summary,
        lead_funnel,
        source_perf,
        industry_distribution,
        country_distribution,
        size_distribution,
        revenue_distribution,
        job_title_distribution,
        leads_by_industry,
    )

    analytics_path = _csv_path("business_analytics_dataset.csv")
    lead_funnel_path = _csv_path("lead_funnel_summary.csv")
    source_perf_path = _csv_path("lead_source_performance.csv")
    industry_path = _csv_path("industry_distribution.csv")
    country_path = _csv_path("country_distribution.csv")

    analytics_df.to_csv(analytics_path, index=False)
    lead_funnel.to_csv(lead_funnel_path, index=False)
    source_perf.to_csv(source_perf_path, index=False)
    industry_distribution.to_csv(industry_path, index=False)
    country_distribution.to_csv(country_path, index=False)

    report_path = write_business_analytics_report(
        executive_summary,
        lead_funnel,
        source_perf,
        industry_distribution,
        country_distribution,
        size_distribution,
        revenue_distribution,
        job_title_distribution,
        accounts,
        contacts,
        leads,
    )

    log.info("-" * 70)
    log.info("BUSINESS KPI SUMMARY")
    for _, row in executive_summary.iterrows():
        log.info("  %-24s %s", row["metric_name"], row["metric_value"])
    log.info("=" * 70)
    log.info("MODULE 8 COMPLETE")
    log.info("=" * 70)

    return {
        "business_analytics_dataset": analytics_df,
        "lead_funnel": lead_funnel,
        "lead_source_performance": source_perf,
        "industry_distribution": industry_distribution,
        "country_distribution": country_distribution,
        "business_analytics_report": report_path,
        "output_paths": {
            "business_analytics_dataset": analytics_path,
            "lead_funnel_summary": lead_funnel_path,
            "lead_source_performance": source_perf_path,
            "industry_distribution": industry_path,
            "country_distribution": country_path,
            "business_analytics_report": report_path,
            "module8_log": os.path.join(LOGS_DIR, "module8_business_analytics.log"),
        },
    }


if __name__ == "__main__":
    run_business_analytics()
