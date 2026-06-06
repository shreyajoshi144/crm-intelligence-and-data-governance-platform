"""
═══════════════════════════════════════════════════════════════════════════════
MODULE 3 — ENTITY RESOLUTION ENGINE
Enterprise CRM Governance & Analytics Platform
═══════════════════════════════════════════════════════════════════════════════

PURPOSE:
    Identify records across tables that represent the same real-world entity.
    Produces entity_match_table — the backbone that Module 4 uses to deduplicate.

MATCHING STRATEGY:
    Pass 1 — EXACT matching
        Accounts : same match_key_name  (normalised company name)
        Contacts : same match_key_email (lowercased email)
        Leads    : same match_key_name  (normalised lead name)

    Pass 2 — FUZZY matching  (SequenceMatcher ratio ≥ FUZZY_THRESHOLD_MEDIUM)
        Accounts : fuzzy on match_key_name for records not matched in Pass 1
        Contacts : fuzzy on match_key_name for unmatched contacts

    Pass 3 — GROUND TRUTH from Golden Record Map
        Uses std_golden_map as pre-confirmed entity links (confidence = 1.0).

    Every match produces one row in entity_match_table with:
        source_record_id   — ID of the candidate record
        matched_record_id  — ID it matched to (canonical / lower ID)
        match_type         — EXACT | FUZZY | GOLDEN
        match_score        — 0.0–1.0
        confidence_level   — HIGH | MEDIUM | LOW
        entity_type        — ACCOUNT | CONTACT | LEAD
        match_field        — column used for matching

OUTPUTS:
    data/processed/entity_match_table.csv
    reports/entity_resolution_report.xlsx
    logs/module3_entity_resolution.log
"""

import os
import re
import sys
import logging
import warnings
from difflib import SequenceMatcher
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings("ignore")

# ── Make the project root importable ──────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config.config import PROCESSED_DIR, LOGS_DIR, REPORTS_DIR

# ── Logger ────────────────────────────────────────────────────────────────────
os.makedirs(LOGS_DIR, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(
            os.path.join(LOGS_DIR, "module3_entity_resolution.log"), mode="w"
        ),
    ],
)
log = logging.getLogger(__name__)

# ── Thresholds ────────────────────────────────────────────────────────────────
FUZZY_THRESHOLD_HIGH   = 0.90   # HIGH confidence
FUZZY_THRESHOLD_MEDIUM = 0.75   # lower bound for inclusion


# ─────────────────────────────────────────────────────────────────────────────
# HELPER UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

def _confidence(score: float, match_type: str) -> str:
    if match_type in ("EXACT", "GOLDEN"):
        return "HIGH"
    if score >= FUZZY_THRESHOLD_HIGH:
        return "HIGH"
    if score >= FUZZY_THRESHOLD_MEDIUM:
        return "MEDIUM"
    return "LOW"


def _fuzzy_ratio(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


def _make_match_key_name(value) -> str:
    """
    Normalise a company / lead name for matching:
        - lowercase
        - strip common legal suffixes (pvt, ltd, llc, inc, corp, plc, limited)
        - remove punctuation
        - collapse whitespace
    """
    if pd.isna(value) or not isinstance(value, str):
        return ""
    v = value.lower().strip()
    # Remove common legal suffixes
    suffixes = [
        r"\bprivate limited\b", r"\bpvt\.?\s*ltd\.?\b", r"\bpvt\b",
        r"\bltd\.?\b", r"\blimited\b", r"\bllc\.?\b", r"\binc\.?\b",
        r"\bcorp\.?\b", r"\bplc\.?\b", r"\bco\.?\b",
    ]
    for suffix in suffixes:
        v = re.sub(suffix, "", v)
    # Remove punctuation and collapse whitespace
    v = re.sub(r"[^a-z0-9\s]", " ", v)
    v = re.sub(r"\s+", " ", v).strip()
    return v


def _make_match_key_email(value) -> str:
    """Lowercased, stripped email for exact matching."""
    if pd.isna(value) or not isinstance(value, str):
        return ""
    return value.strip().lower()


def _add_match_keys_accounts(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["match_key_name"] = out["account_name"].apply(_make_match_key_name)
    return out


def _add_match_keys_contacts(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["match_key_email"] = out["email"].apply(_make_match_key_email)
    out["match_key_name"]  = (out["first_name"].fillna("") + " " +
                               out["last_name"].fillna("")).apply(_make_match_key_name)
    return out


def _add_match_keys_leads(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["match_key_name"] = out["lead_name"].apply(_make_match_key_name)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# PASS 1 — EXACT MATCHING
# ─────────────────────────────────────────────────────────────────────────────

def exact_match_on_key(
    df: pd.DataFrame,
    id_col: str,
    key_col: str,
    entity_type: str,
) -> tuple[pd.DataFrame, set]:
    """
    Group records by key_col.  Any group with ≥ 2 members is a match cluster.
    Within each cluster emit (non-canonical → canonical) pairs where the
    canonical is the lexicographically smallest ID.

    Returns:
        (match_df, matched_id_set)
    """
    log.info(f"    Pass 1 — EXACT on '{key_col}' for {entity_type}")
    matched_ids: set = set()
    rows: list[dict] = []

    valid = df[df[key_col].notna() & (df[key_col] != "")].copy()
    groups = valid.groupby(key_col)[id_col].apply(list)

    for key, members in groups.items():
        if len(members) < 2:
            continue
        members_sorted = sorted(members)
        canonical = members_sorted[0]
        for other in members_sorted[1:]:
            rows.append({
                "source_record_id":  other,
                "matched_record_id": canonical,
                "entity_type":       entity_type,
                "match_type":        "EXACT",
                "match_field":       key_col,
                "match_score":       1.0,
                "confidence_level":  "HIGH",
            })
            matched_ids.add(other)
            matched_ids.add(canonical)

    match_df = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=["source_record_id", "matched_record_id", "entity_type",
                 "match_type", "match_field", "match_score", "confidence_level"]
    )
    log.info(f"      Exact matches found : {len(rows)}  "
             f"({len(matched_ids)} unique IDs involved)")
    return match_df, matched_ids


# ─────────────────────────────────────────────────────────────────────────────
# PASS 2 — FUZZY MATCHING
# ─────────────────────────────────────────────────────────────────────────────

def fuzzy_match_on_key(
    df: pd.DataFrame,
    id_col: str,
    key_col: str,
    entity_type: str,
    already_matched: set,
    sample_size: int = 3000,
) -> pd.DataFrame:
    """
    Sliding-window fuzzy match on records NOT already captured in Pass 1.

    Alphabetically sorted → compared against the next WINDOW neighbours.
    This is the standard production MDM approach (records that are
    alphabetically adjacent are most likely to be fuzzy-equivalent).

    Returns match_df.
    """
    WINDOW = 10
    log.info(f"    Pass 2 — FUZZY on '{key_col}' for {entity_type}  (window={WINDOW})")

    unmatched = df[
        df[key_col].notna()
        & (df[key_col] != "")
        & ~df[id_col].isin(already_matched)
    ].copy()

    if len(unmatched) > sample_size:
        unmatched = unmatched.sample(n=sample_size, random_state=42)
        log.info(f"      Sampled {sample_size} unmatched records for fuzzy pass")

    unmatched = unmatched.sort_values(key_col).reset_index(drop=True)
    records   = list(zip(unmatched[id_col], unmatched[key_col]))
    rows: list[dict] = []

    for i, (id_a, key_a) in enumerate(records):
        for j in range(i + 1, min(i + 1 + WINDOW, len(records))):
            id_b, key_b = records[j]
            score = _fuzzy_ratio(key_a, key_b)
            if score >= FUZZY_THRESHOLD_MEDIUM:
                rows.append({
                    "source_record_id":  id_b,
                    "matched_record_id": id_a,
                    "entity_type":       entity_type,
                    "match_type":        "FUZZY",
                    "match_field":       key_col,
                    "match_score":       round(score, 4),
                    "confidence_level":  _confidence(score, "FUZZY"),
                })

    result = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=["source_record_id", "matched_record_id", "entity_type",
                 "match_type", "match_field", "match_score", "confidence_level"]
    )
    log.info(f"      Fuzzy matches found : {len(rows)}")
    return result


# ─────────────────────────────────────────────────────────────────────────────
# PASS 3 — GOLDEN RECORD GROUND TRUTH
# ─────────────────────────────────────────────────────────────────────────────

def golden_record_matches(golden_df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert std_golden_map into entity_match_table rows.
    Accounts sharing the same golden_record_id are the same entity.
    Canonical = lexicographically first account_id in each golden group.
    """
    log.info("    Pass 3 — GOLDEN RECORD ground-truth links")
    rows: list[dict] = []

    groups = golden_df.groupby("golden_record_id")["account_id"].apply(list)
    for gold_id, members in groups.items():
        if len(members) < 2:
            continue
        members_sorted = sorted(members)
        canonical = members_sorted[0]
        for other in members_sorted[1:]:
            rows.append({
                "source_record_id":  other,
                "matched_record_id": canonical,
                "entity_type":       "ACCOUNT",
                "match_type":        "GOLDEN",
                "match_field":       "golden_record_id",
                "match_score":       1.0,
                "confidence_level":  "HIGH",
            })

    result = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=["source_record_id", "matched_record_id", "entity_type",
                 "match_type", "match_field", "match_score", "confidence_level"]
    )
    log.info(f"      Golden matches : {len(rows)}")
    return result


# ─────────────────────────────────────────────────────────────────────────────
# DEDUPLICATE THE MATCH TABLE (remove redundant pairs)
# ─────────────────────────────────────────────────────────────────────────────

def deduplicate_match_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    A pair (A→B) can appear from EXACT and also from GOLDEN — keep only the
    highest-priority one.
    Priority order: GOLDEN > EXACT > FUZZY
    """
    if len(df) == 0:
        return df
    priority = {"GOLDEN": 0, "EXACT": 1, "FUZZY": 2}
    df = df.copy()
    df["_priority"] = df["match_type"].map(priority)
    df = df.sort_values("_priority")
    df = df.drop_duplicates(
        subset=["source_record_id", "matched_record_id"], keep="first"
    )
    df = df.drop(columns=["_priority"]).reset_index(drop=True)
    return df


# ─────────────────────────────────────────────────────────────────────────────
# REPORT
# ─────────────────────────────────────────────────────────────────────────────

def write_entity_resolution_report(match_df: pd.DataFrame) -> str:
    path = os.path.join(REPORTS_DIR, "entity_resolution_report.xlsx")

    HDR_FILL  = PatternFill("solid", start_color="1F3864")
    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    BODY_FONT = Font(name="Arial", size=10)
    HIGH_FILL = PatternFill("solid", start_color="C6EFCE")
    HIGH_FONT = Font(bold=True, color="276221", name="Arial", size=10)
    MED_FILL  = PatternFill("solid", start_color="FFEB9C")
    MED_FONT  = Font(bold=True, color="9C5700", name="Arial", size=10)
    LOW_FILL  = PatternFill("solid", start_color="FFC7CE")
    LOW_FONT  = Font(bold=True, color="9C0006", name="Arial", size=10)
    CENTER    = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="CCCCCC")
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.freeze_panes = "A2"

    summary_rows = []
    for entity_type, grp in match_df.groupby("entity_type"):
        for match_type, grp2 in grp.groupby("match_type"):
            summary_rows.append({
                "Entity Type":       entity_type,
                "Match Type":        match_type,
                "Total Matches":     len(grp2),
                "HIGH Confidence":   int((grp2["confidence_level"] == "HIGH").sum()),
                "MEDIUM Confidence": int((grp2["confidence_level"] == "MEDIUM").sum()),
                "LOW Confidence":    int((grp2["confidence_level"] == "LOW").sum()),
                "Avg Score":         round(grp2["match_score"].mean(), 4),
            })

    if len(match_df) > 0:
        summary_rows.append({
            "Entity Type":       "TOTAL",
            "Match Type":        "ALL",
            "Total Matches":     len(match_df),
            "HIGH Confidence":   int((match_df["confidence_level"] == "HIGH").sum()),
            "MEDIUM Confidence": int((match_df["confidence_level"] == "MEDIUM").sum()),
            "LOW Confidence":    int((match_df["confidence_level"] == "LOW").sum()),
            "Avg Score":         round(match_df["match_score"].mean(), 4),
        })

    h1 = ["Entity Type", "Match Type", "Total Matches",
          "HIGH Confidence", "MEDIUM Confidence", "LOW Confidence", "Avg Score"]
    ws1.append(h1)
    for cell in ws1[1]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for row in summary_rows:
        ws1.append([row[h] for h in h1])
        r = ws1.max_row
        for c in ws1[r]:
            c.font = BODY_FONT
            c.alignment = CENTER
            c.border = BORDER

    for col, w in zip(["A","B","C","D","E","F","G"], [14, 10, 14, 16, 18, 14, 10]):
        ws1.column_dimensions[col].width = w

    # ── Sheet 2: Full match table (first 5 000 rows) ───────────────────────
    ws2 = wb.create_sheet("Match Table")
    ws2.freeze_panes = "A2"

    h2 = ["source_record_id", "matched_record_id", "entity_type",
          "match_type", "match_field", "match_score", "confidence_level"]
    ws2.append(h2)
    for cell in ws2[1]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for _, row in match_df.head(5000).iterrows():
        ws2.append([row[c] for c in h2])
        r = ws2.max_row
        conf_cell = ws2.cell(row=r, column=7)
        if conf_cell.value == "HIGH":
            conf_cell.fill = HIGH_FILL
            conf_cell.font = HIGH_FONT
        elif conf_cell.value == "MEDIUM":
            conf_cell.fill = MED_FILL
            conf_cell.font = MED_FONT
        else:
            conf_cell.fill = LOW_FILL
            conf_cell.font = LOW_FONT
        for c in ws2[r]:
            if c.column != 7:
                c.font = BODY_FONT
            c.alignment = CENTER
            c.border = BORDER

    for col, w in zip(["A","B","C","D","E","F","G"], [16, 16, 12, 10, 20, 12, 16]):
        ws2.column_dimensions[col].width = w

    # ── Sheet 3: Run info ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Run Info")
    ws3["A1"] = "Generated"
    ws3["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws3["A2"] = "Module"
    ws3["B2"] = "Module 3 — Entity Resolution Engine"
    ws3["A3"] = "Platform"
    ws3["B3"] = "Enterprise CRM Governance & Analytics Platform"
    ws3["A4"] = "Total Matches"
    ws3["B4"] = len(match_df)
    ws3["A5"] = "Fuzzy Threshold HIGH"
    ws3["B5"] = FUZZY_THRESHOLD_HIGH
    ws3["A6"] = "Fuzzy Threshold MEDIUM"
    ws3["B6"] = FUZZY_THRESHOLD_MEDIUM
    for row in ws3.iter_rows(min_row=1, max_row=6, min_col=1, max_col=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10)

    wb.save(path)
    log.info(f"  Entity resolution report → {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ORCHESTRATOR
# ─────────────────────────────────────────────────────────────────────────────

def run_entity_resolution(std_data: dict) -> pd.DataFrame:
    """
    Args:
        std_data : dict returned by module2_standardization.run_standardization()

    Returns:
        entity_match_table  (DataFrame)
    """
    log.info("=" * 70)
    log.info("MODULE 3 — ENTITY RESOLUTION ENGINE")
    log.info("=" * 70)

    os.makedirs(PROCESSED_DIR, exist_ok=True)

    std_accounts = std_data.get("std_accounts", pd.DataFrame())
    std_contacts = std_data.get("std_contacts", pd.DataFrame())
    std_leads    = std_data.get("std_leads",    pd.DataFrame())
    std_golden   = std_data.get("std_golden_map", pd.DataFrame())

    # ── Add match keys ─────────────────────────────────────────────────────
    if len(std_accounts) > 0:
        std_accounts = _add_match_keys_accounts(std_accounts)
    if len(std_contacts) > 0:
        std_contacts = _add_match_keys_contacts(std_contacts)
    if len(std_leads) > 0:
        std_leads = _add_match_keys_leads(std_leads)

    all_match_rows: list[pd.DataFrame] = []

    # ── ACCOUNTS ──────────────────────────────────────────────────────────
    log.info("  ─── Resolving: ACCOUNTS")
    if len(std_accounts) > 0:
        acc_exact_df, acc_exact_ids = exact_match_on_key(
            std_accounts, "account_id", "match_key_name", "ACCOUNT"
        )
        acc_fuzzy_df = fuzzy_match_on_key(
            std_accounts, "account_id", "match_key_name", "ACCOUNT",
            already_matched=acc_exact_ids,
        )
        all_match_rows.extend([acc_exact_df, acc_fuzzy_df])
    else:
        log.warning("  ⚠ std_accounts is empty — skipping account resolution")

    # ── CONTACTS ──────────────────────────────────────────────────────────
    log.info("  ─── Resolving: CONTACTS")
    if len(std_contacts) > 0:
        con_email_df, con_exact_ids = exact_match_on_key(
            std_contacts, "contact_id", "match_key_email", "CONTACT"
        )
        con_fuzzy_df = fuzzy_match_on_key(
            std_contacts, "contact_id", "match_key_name", "CONTACT",
            already_matched=con_exact_ids,
        )
        all_match_rows.extend([con_email_df, con_fuzzy_df])
    else:
        log.warning("  ⚠ std_contacts is empty — skipping contact resolution")

    # ── LEADS ─────────────────────────────────────────────────────────────
    log.info("  ─── Resolving: LEADS")
    if len(std_leads) > 0:
        lead_exact_df, _ = exact_match_on_key(
            std_leads, "lead_id", "match_key_name", "LEAD"
        )
        all_match_rows.append(lead_exact_df)
    else:
        log.warning("  ⚠ std_leads is empty — skipping lead resolution")

    # ── GOLDEN GROUND TRUTH ───────────────────────────────────────────────
    if std_golden is not None and len(std_golden) > 0:
        log.info("  ─── Golden Record ground-truth pass")
        golden_matches = golden_record_matches(std_golden)
        all_match_rows.append(golden_matches)
    else:
        log.info("  ─── No golden map available — skipping Pass 3")

    # ── Assemble + deduplicate ─────────────────────────────────────────────
    non_empty = [df for df in all_match_rows if df is not None and len(df) > 0]
    if non_empty:
        entity_match_table = pd.concat(non_empty, ignore_index=True)
        entity_match_table = deduplicate_match_table(entity_match_table)
    else:
        entity_match_table = pd.DataFrame(
            columns=["source_record_id", "matched_record_id", "entity_type",
                     "match_type", "match_field", "match_score", "confidence_level"]
        )

    # ── Save ──────────────────────────────────────────────────────────────
    out_path = os.path.join(PROCESSED_DIR, "entity_match_table.csv")
    entity_match_table.to_csv(out_path, index=False)
    log.info(f"  entity_match_table → {out_path}  ({len(entity_match_table):,} rows)")

    # ── Summary ───────────────────────────────────────────────────────────
    log.info("-" * 70)
    if len(entity_match_table) > 0:
        for (et, mt), grp in entity_match_table.groupby(["entity_type", "match_type"]):
            log.info(f"    {et:<10} {mt:<8} → {len(grp):>6,} matches")
    else:
        log.info("    No entity matches found.")
    log.info("-" * 70)

    write_entity_resolution_report(entity_match_table)

    log.info("=" * 70)
    log.info(f"MODULE 3 COMPLETE — {len(entity_match_table):,} entity matches identified")
    log.info("=" * 70)

    return entity_match_table


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    from module1_ingestion import run_ingestion
    from module2_standardization import run_standardization
    staging = run_ingestion()
    std     = run_standardization(staging)
    run_entity_resolution(std)
