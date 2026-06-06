"""
MODULE 5 - DATA QUALITY ENGINE
Enterprise CRM Governance & Analytics Platform

Purpose
-------
Measures CRM health after deduplication and produces:
    data/processed/dq_scores.csv
    data/processed/dq_issue_log.csv
    data/processed/quality_scores_accounts.csv
    data/processed/quality_scores_contacts.csv
    data/processed/quality_scores_leads.csv
    reports/data_quality_report.xlsx
    logs/module5_data_quality.log

Public entry point:
    run_data_quality(dedup_data: dict, std_data: dict) -> dict

The return payload is Module 6 friendly. It includes aggregate scores, issue
logs, and record-level quality score DataFrames that can be merged into golden
record selection logic.
"""

from __future__ import annotations

import logging
import os
import re
import sys
import warnings
from datetime import date, datetime
from typing import Iterable

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

warnings.filterwarnings("ignore")

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from config.config import (
        COUNTRY_MAP,
        INDUSTRY_MAP,
        LOGS_DIR,
        PROCESSED_DIR,
        REPORTS_DIR,
        VALID_LEAD_SOURCES,
        VALID_LEAD_STATUSES,
    )
except Exception:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    PROCESSED_DIR = os.path.join(BASE_DIR, "data", "processed")
    LOGS_DIR = os.path.join(BASE_DIR, "logs")
    REPORTS_DIR = os.path.join(BASE_DIR, "reports")
    INDUSTRY_MAP = {}
    COUNTRY_MAP = {}
    VALID_LEAD_SOURCES = [
        "Web", "Referral", "Event", "LinkedIn", "Cold Call",
        "Partner", "Advertisement", "Other",
    ]
    VALID_LEAD_STATUSES = [
        "New", "Working", "Qualified", "Converted", "Unqualified", "Nurturing",
    ]


for directory in (PROCESSED_DIR, LOGS_DIR, REPORTS_DIR):
    os.makedirs(directory, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(os.path.join(LOGS_DIR, "module5_data_quality.log"), mode="w"),
    ],
)
log = logging.getLogger(__name__)


WEIGHTS = {
    "completeness": 0.35,
    "validity": 0.25,
    "consistency": 0.20,
    "uniqueness": 0.10,
    "freshness": 0.10,
}

EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")
PHONE_RE = re.compile(r"^\d{7,15}$")
ACCID_RE = re.compile(r"^ACC\d{6}$")
CONID_RE = re.compile(r"^CON\d{6}$")
LEDID_RE = re.compile(r"^LED\d{6}$")

KNOWN_INDUSTRIES = set(INDUSTRY_MAP.values())
KNOWN_COUNTRIES = set(COUNTRY_MAP.values())
FRESHNESS_DAYS = 730
TODAY = date.today()

ISSUE_COLUMNS = [
    "record_id", "entity_type", "dimension", "dq_check",
    "field_name", "issue_value", "severity",
]


def _pct(passing: int, total: int) -> float:
    if total == 0:
        return 100.0
    return round((passing / total) * 100, 2)


def _blank_mask(df: pd.DataFrame, col: str) -> pd.Series:
    if col not in df.columns:
        return pd.Series([True] * len(df), index=df.index)
    return df[col].isna() | df[col].astype(str).str.strip().eq("")


def _clean_str(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _empty_issues() -> pd.DataFrame:
    return pd.DataFrame(columns=ISSUE_COLUMNS)


def _flag(
    df: pd.DataFrame,
    mask: pd.Series,
    check_name: str,
    entity_type: str,
    id_col: str,
    field_name: str = "",
    severity: str = "MEDIUM",
) -> pd.DataFrame:
    if len(df) == 0:
        return _empty_issues()

    mask = mask.reindex(df.index, fill_value=False).fillna(False)
    failing = df.loc[mask].copy()
    if len(failing) == 0:
        return _empty_issues()

    if id_col in failing.columns:
        record_ids = failing[id_col].astype(str)
    else:
        record_ids = failing.index.astype(str)

    if field_name and field_name in failing.columns:
        issue_values = failing[field_name].map(_clean_str)
    else:
        issue_values = ""

    return pd.DataFrame({
        "record_id": record_ids,
        "entity_type": entity_type,
        "dimension": check_name.split("_", 1)[0].capitalize(),
        "dq_check": check_name,
        "field_name": field_name,
        "issue_value": issue_values,
        "severity": severity,
    })


def _concat_issues(issue_frames: Iterable[pd.DataFrame]) -> pd.DataFrame:
    non_empty = [df for df in issue_frames if df is not None and len(df) > 0]
    if not non_empty:
        return _empty_issues()
    return pd.concat(non_empty, ignore_index=True)[ISSUE_COLUMNS]


def _weighted_entity_score(scores: dict, dimension: str, counts: dict[str, int]) -> float:
    total = sum(counts.values())
    if total == 0:
        return 100.0
    value = (
        scores.get(f"accounts_{dimension}", 100.0) * counts["accounts"]
        + scores.get(f"contacts_{dimension}", 100.0) * counts["contacts"]
        + scores.get(f"leads_{dimension}", 100.0) * counts["leads"]
    ) / total
    return round(value, 2)


def check_completeness(
    accounts: pd.DataFrame,
    contacts: pd.DataFrame,
    leads: pd.DataFrame,
) -> tuple[dict, pd.DataFrame]:
    log.info("    Completeness checks")
    issues = []
    scores = {}

    required = {
        "accounts": ("ACCOUNT", accounts, "account_id", [
            "account_name", "industry", "country", "annual_revenue", "employee_count",
        ]),
        "contacts": ("CONTACT", contacts, "contact_id", [
            "first_name", "last_name", "email", "job_title", "account_id",
        ]),
        "leads": ("LEAD", leads, "lead_id", ["lead_name", "source", "status"]),
    }

    for entity_key, (entity_type, df, id_col, columns) in required.items():
        if len(df) == 0:
            scores[f"{entity_key}_completeness"] = 100.0
            continue
        failing_any = pd.Series([False] * len(df), index=df.index)
        for col in columns:
            mask = _blank_mask(df, col)
            failing_any = failing_any | mask
            issues.append(_flag(
                df, mask, f"completeness_missing_{col}",
                entity_type, id_col, col, "HIGH",
            ))
        scores[f"{entity_key}_completeness"] = _pct(int((~failing_any).sum()), len(df))
        log.info(
            "      %-8s completeness: %.1f%%",
            entity_type.title(), scores[f"{entity_key}_completeness"],
        )

    counts = {"accounts": len(accounts), "contacts": len(contacts), "leads": len(leads)}
    scores["overall_completeness"] = _weighted_entity_score(scores, "completeness", counts)
    return scores, _concat_issues(issues)


def check_validity(
    accounts: pd.DataFrame,
    contacts: pd.DataFrame,
    leads: pd.DataFrame,
) -> tuple[dict, pd.DataFrame]:
    log.info("    Validity checks")
    issues = []
    scores = {}

    if len(accounts) > 0:
        failing = pd.Series([False] * len(accounts), index=accounts.index)

        if "industry" in accounts.columns and KNOWN_INDUSTRIES:
            mask = ~accounts["industry"].isin(KNOWN_INDUSTRIES)
            failing = failing | mask
            issues.append(_flag(accounts, mask, "validity_invalid_industry", "ACCOUNT", "account_id", "industry", "MEDIUM"))

        if "country" in accounts.columns and KNOWN_COUNTRIES:
            mask = ~accounts["country"].isin(KNOWN_COUNTRIES)
            failing = failing | mask
            issues.append(_flag(accounts, mask, "validity_invalid_country", "ACCOUNT", "account_id", "country", "MEDIUM"))

        if "annual_revenue" in accounts.columns:
            revenue = pd.to_numeric(accounts["annual_revenue"], errors="coerce")
            mask = accounts["annual_revenue"].notna() & revenue.isna()
            failing = failing | mask
            issues.append(_flag(accounts, mask, "validity_non_numeric_annual_revenue", "ACCOUNT", "account_id", "annual_revenue", "MEDIUM"))
            mask = revenue.notna() & (revenue < 0)
            failing = failing | mask
            issues.append(_flag(accounts, mask, "validity_negative_annual_revenue", "ACCOUNT", "account_id", "annual_revenue", "HIGH"))

        if "employee_count" in accounts.columns:
            employees = pd.to_numeric(accounts["employee_count"], errors="coerce")
            mask = accounts["employee_count"].notna() & employees.isna()
            failing = failing | mask
            issues.append(_flag(accounts, mask, "validity_non_numeric_employee_count", "ACCOUNT", "account_id", "employee_count", "MEDIUM"))
            mask = employees.notna() & (employees < 0)
            failing = failing | mask
            issues.append(_flag(accounts, mask, "validity_negative_employee_count", "ACCOUNT", "account_id", "employee_count", "HIGH"))

        scores["accounts_validity"] = _pct(int((~failing).sum()), len(accounts))
    else:
        scores["accounts_validity"] = 100.0

    if len(contacts) > 0:
        failing = pd.Series([False] * len(contacts), index=contacts.index)

        if "email" in contacts.columns:
            mask = contacts["email"].apply(lambda v: not bool(EMAIL_RE.match(_clean_str(v))))
            if "email_valid" in contacts.columns:
                mask = mask | contacts["email_valid"].astype(str).str.upper().eq("NO")
            failing = failing | mask
            issues.append(_flag(contacts, mask, "validity_invalid_email", "CONTACT", "contact_id", "email", "HIGH"))

        if "account_id" in contacts.columns:
            mask = contacts["account_id"].notna() & ~contacts["account_id"].astype(str).str.match(ACCID_RE)
            failing = failing | mask
            issues.append(_flag(contacts, mask, "validity_invalid_account_id_format", "CONTACT", "contact_id", "account_id", "MEDIUM"))

        if "contact_id" in contacts.columns:
            mask = contacts["contact_id"].notna() & ~contacts["contact_id"].astype(str).str.match(CONID_RE)
            failing = failing | mask
            issues.append(_flag(contacts, mask, "validity_invalid_contact_id_format", "CONTACT", "contact_id", "contact_id", "MEDIUM"))

        if "phone" in contacts.columns:
            mask = contacts["phone"].notna() & contacts["phone"].astype(str).str.strip().ne("") & ~contacts["phone"].astype(str).str.match(PHONE_RE)
            failing = failing | mask
            issues.append(_flag(contacts, mask, "validity_invalid_phone", "CONTACT", "contact_id", "phone", "LOW"))

        scores["contacts_validity"] = _pct(int((~failing).sum()), len(contacts))
    else:
        scores["contacts_validity"] = 100.0

    if len(leads) > 0:
        failing = pd.Series([False] * len(leads), index=leads.index)

        if "status" in leads.columns:
            mask = ~leads["status"].isin(VALID_LEAD_STATUSES)
            failing = failing | mask
            issues.append(_flag(leads, mask, "validity_invalid_lead_status", "LEAD", "lead_id", "status", "MEDIUM"))

        if "source" in leads.columns:
            mask = ~leads["source"].isin(VALID_LEAD_SOURCES)
            failing = failing | mask
            issues.append(_flag(leads, mask, "validity_invalid_lead_source", "LEAD", "lead_id", "source", "MEDIUM"))

        if "lead_id" in leads.columns:
            mask = leads["lead_id"].notna() & ~leads["lead_id"].astype(str).str.match(LEDID_RE)
            failing = failing | mask
            issues.append(_flag(leads, mask, "validity_invalid_lead_id_format", "LEAD", "lead_id", "lead_id", "MEDIUM"))

        scores["leads_validity"] = _pct(int((~failing).sum()), len(leads))
    else:
        scores["leads_validity"] = 100.0

    counts = {"accounts": len(accounts), "contacts": len(contacts), "leads": len(leads)}
    scores["overall_validity"] = _weighted_entity_score(scores, "validity", counts)
    log.info("      Overall validity: %.1f%%", scores["overall_validity"])
    return scores, _concat_issues(issues)


def check_consistency(
    accounts: pd.DataFrame,
    contacts: pd.DataFrame,
    leads: pd.DataFrame,
    account_dedup_map: pd.DataFrame | None = None,
) -> tuple[dict, pd.DataFrame]:
    log.info("    Consistency checks")
    issues = []
    scores = {}

    known_account_ids = set(accounts.get("account_id", pd.Series(dtype=str)).dropna().astype(str))
    if account_dedup_map is not None and len(account_dedup_map) > 0:
        for col in ("record_id", "canonical_id"):
            if col in account_dedup_map.columns:
                known_account_ids.update(account_dedup_map[col].dropna().astype(str))

    if len(accounts) > 0:
        failing = pd.Series([False] * len(accounts), index=accounts.index)
        if {"annual_revenue", "employee_count"}.issubset(accounts.columns):
            revenue = pd.to_numeric(accounts["annual_revenue"], errors="coerce")
            employees = pd.to_numeric(accounts["employee_count"], errors="coerce")
            mask = (revenue > 0) & (employees == 0)
            failing = failing | mask
            issues.append(_flag(accounts, mask, "consistency_revenue_without_employees", "ACCOUNT", "account_id", "employee_count", "LOW"))
        scores["accounts_consistency"] = _pct(int((~failing).sum()), len(accounts))
    else:
        scores["accounts_consistency"] = 100.0

    if len(contacts) > 0:
        failing = pd.Series([False] * len(contacts), index=contacts.index)
        if "account_id" in contacts.columns and known_account_ids:
            mask = contacts["account_id"].notna() & ~contacts["account_id"].astype(str).isin(known_account_ids)
            failing = failing | mask
            issues.append(_flag(contacts, mask, "consistency_orphan_contact_account_id", "CONTACT", "contact_id", "account_id", "HIGH"))
        scores["contacts_consistency"] = _pct(int((~failing).sum()), len(contacts))
    else:
        scores["contacts_consistency"] = 100.0

    if len(leads) > 0:
        failing = pd.Series([False] * len(leads), index=leads.index)
        if {"status", "source"}.issubset(leads.columns):
            mask = leads["status"].eq("Converted") & _blank_mask(leads, "source")
            failing = failing | mask
            issues.append(_flag(leads, mask, "consistency_converted_lead_missing_source", "LEAD", "lead_id", "source", "MEDIUM"))
        scores["leads_consistency"] = _pct(int((~failing).sum()), len(leads))
    else:
        scores["leads_consistency"] = 100.0

    counts = {"accounts": len(accounts), "contacts": len(contacts), "leads": len(leads)}
    scores["overall_consistency"] = _weighted_entity_score(scores, "consistency", counts)
    return scores, _concat_issues(issues)


def check_uniqueness(
    accounts: pd.DataFrame,
    contacts: pd.DataFrame,
    leads: pd.DataFrame,
    stewardship: pd.DataFrame | None,
) -> tuple[dict, pd.DataFrame]:
    log.info("    Uniqueness checks")
    issues = []
    scores = {}

    specs = [
        ("accounts", "ACCOUNT", accounts, "account_id"),
        ("contacts", "CONTACT", contacts, "contact_id"),
        ("leads", "LEAD", leads, "lead_id"),
    ]

    for entity_key, entity_type, df, id_col in specs:
        if len(df) == 0:
            scores[f"{entity_key}_uniqueness"] = 100.0
            continue

        failing = pd.Series([False] * len(df), index=df.index)
        if id_col in df.columns:
            mask = df.duplicated(subset=[id_col], keep=False)
            failing = failing | mask
            issues.append(_flag(df, mask, f"uniqueness_duplicate_{id_col}", entity_type, id_col, id_col, "HIGH"))

        if entity_key == "accounts" and stewardship is not None and len(stewardship) > 0:
            if {"account_id", "dq_issue_status"}.issubset(stewardship.columns):
                open_mask = stewardship["dq_issue_status"].isin(["Open", "In Review", "Escalated"])
                open_accounts = set(stewardship.loc[open_mask, "account_id"].dropna().astype(str))
                mask = df["account_id"].astype(str).isin(open_accounts) if "account_id" in df.columns else pd.Series(False, index=df.index)
                failing = failing | mask
                issues.append(_flag(df, mask, "uniqueness_open_stewardship_issue", entity_type, id_col, "account_id", "MEDIUM"))

        if entity_key == "contacts" and "email" in df.columns:
            valid_email = df["email"].notna() & df["email"].astype(str).str.strip().ne("")
            mask = valid_email & df.duplicated(subset=["email"], keep=False)
            failing = failing | mask
            issues.append(_flag(df, mask, "uniqueness_duplicate_email", entity_type, id_col, "email", "HIGH"))

        if entity_key == "leads" and {"lead_name", "source"}.issubset(df.columns):
            valid_name = df["lead_name"].notna() & df["lead_name"].astype(str).str.strip().ne("")
            mask = valid_name & df.duplicated(subset=["lead_name", "source"], keep=False)
            failing = failing | mask
            issues.append(_flag(df, mask, "uniqueness_duplicate_lead_name_source", entity_type, id_col, "lead_name", "MEDIUM"))

        scores[f"{entity_key}_uniqueness"] = _pct(int((~failing).sum()), len(df))

    counts = {"accounts": len(accounts), "contacts": len(contacts), "leads": len(leads)}
    scores["overall_uniqueness"] = _weighted_entity_score(scores, "uniqueness", counts)
    return scores, _concat_issues(issues)


def check_freshness(
    accounts: pd.DataFrame,
    contacts: pd.DataFrame,
    leads: pd.DataFrame,
) -> tuple[dict, pd.DataFrame]:
    log.info("    Freshness checks")
    issues = []
    scores = {}
    date_cols = [
        "last_modified_date", "last_activity_date", "created_date",
        "last_modified", "modified_date", "created",
    ]

    def stale_from_date(df: pd.DataFrame) -> pd.Series | None:
        for col in date_cols:
            if col in df.columns:
                dates = pd.to_datetime(df[col], errors="coerce")
                age_days = (pd.Timestamp(TODAY) - dates).dt.days
                return dates.isna() | (age_days > FRESHNESS_DAYS)
        return None

    if len(accounts) > 0:
        stale = stale_from_date(accounts)
        if stale is None:
            if {"annual_revenue", "employee_count"}.issubset(accounts.columns):
                stale = _blank_mask(accounts, "annual_revenue") & _blank_mask(accounts, "employee_count")
            else:
                stale = pd.Series([False] * len(accounts), index=accounts.index)
        issues.append(_flag(accounts, stale, "freshness_stale_account", "ACCOUNT", "account_id", "", "LOW"))
        scores["accounts_freshness"] = _pct(int((~stale).sum()), len(accounts))
    else:
        scores["accounts_freshness"] = 100.0

    if len(contacts) > 0:
        stale = stale_from_date(contacts)
        if stale is None:
            stale = _blank_mask(contacts, "email") & _blank_mask(contacts, "job_title")
        issues.append(_flag(contacts, stale, "freshness_stale_contact", "CONTACT", "contact_id", "", "LOW"))
        scores["contacts_freshness"] = _pct(int((~stale).sum()), len(contacts))
    else:
        scores["contacts_freshness"] = 100.0

    if len(leads) > 0:
        stale = stale_from_date(leads)
        if stale is None:
            if {"status", "source"}.issubset(leads.columns):
                stale = leads["status"].eq("Unqualified") & leads["source"].isin(["Other", ""])
            else:
                stale = pd.Series([False] * len(leads), index=leads.index)
        issues.append(_flag(leads, stale, "freshness_stale_lead", "LEAD", "lead_id", "", "LOW"))
        scores["leads_freshness"] = _pct(int((~stale).sum()), len(leads))
    else:
        scores["leads_freshness"] = 100.0

    counts = {"accounts": len(accounts), "contacts": len(contacts), "leads": len(leads)}
    scores["overall_freshness"] = _weighted_entity_score(scores, "freshness", counts)
    return scores, _concat_issues(issues)


def compute_crm_health_score(all_scores: dict) -> float:
    return round(sum(
        all_scores.get(f"overall_{dimension}", 100.0) * weight
        for dimension, weight in WEIGHTS.items()
    ), 2)


def build_scores_df(all_scores: dict, crm_health: float) -> pd.DataFrame:
    rows = []
    for entity in ("accounts", "contacts", "leads"):
        for dimension in WEIGHTS:
            rows.append({
                "entity": entity.upper(),
                "dimension": dimension.capitalize(),
                "score": all_scores.get(f"{entity}_{dimension}", 100.0),
                "weight": WEIGHTS[dimension],
            })
    for dimension in WEIGHTS:
        rows.append({
            "entity": "OVERALL",
            "dimension": dimension.capitalize(),
            "score": all_scores.get(f"overall_{dimension}", 100.0),
            "weight": WEIGHTS[dimension],
        })
    rows.append({
        "entity": "CRM HEALTH SCORE",
        "dimension": "Weighted Average",
        "score": crm_health,
        "weight": 1.0,
    })
    return pd.DataFrame(rows)


def _grade(score: float) -> str:
    if score >= 90:
        return "A"
    if score >= 80:
        return "B"
    if score >= 70:
        return "C"
    return "D"


def build_record_quality_scores(
    df: pd.DataFrame,
    entity_type: str,
    id_col: str,
    issue_df: pd.DataFrame,
) -> pd.DataFrame:
    columns = [
        id_col, "entity_type", "completeness_score", "validity_score",
        "consistency_score", "uniqueness_score", "freshness_score",
        "record_quality_score", "quality_grade", "issue_count",
        "high_severity_issues", "module6_golden_record_eligible",
    ]

    if len(df) == 0:
        return pd.DataFrame(columns=columns)

    out = pd.DataFrame({
        id_col: df[id_col].astype(str) if id_col in df.columns else df.index.astype(str),
        "entity_type": entity_type,
    })

    relevant = issue_df[issue_df["entity_type"].eq(entity_type)].copy() if len(issue_df) > 0 else _empty_issues()
    issue_groups = {
        rid: grp for rid, grp in relevant.groupby("record_id")
    } if len(relevant) > 0 else {}

    dimension_cols = []
    for dimension in WEIGHTS:
        col = f"{dimension}_score"
        dimension_cols.append(col)
        out[col] = out[id_col].map(
            lambda rid, d=dimension: 0.0
            if rid in issue_groups and issue_groups[rid]["dimension"].str.lower().eq(d).any()
            else 100.0
        )

    out["record_quality_score"] = round(sum(
        out[f"{dimension}_score"] * weight
        for dimension, weight in WEIGHTS.items()
    ), 2)
    out["quality_grade"] = out["record_quality_score"].apply(_grade)
    out["issue_count"] = out[id_col].map(lambda rid: len(issue_groups.get(rid, []))).fillna(0).astype(int)
    out["high_severity_issues"] = out[id_col].map(
        lambda rid: int((issue_groups[rid]["severity"] == "HIGH").sum()) if rid in issue_groups else 0
    )
    out["module6_golden_record_eligible"] = out.apply(
        lambda row: "YES" if row["record_quality_score"] >= 80 and row["high_severity_issues"] == 0 else "NO",
        axis=1,
    )
    return out[columns]


def write_data_quality_report(
    scores_df: pd.DataFrame,
    issue_df: pd.DataFrame,
    record_scores: dict[str, pd.DataFrame],
    crm_health: float,
    all_scores: dict,
) -> str:
    path = os.path.join(REPORTS_DIR, "data_quality_report.xlsx")

    navy = "1F3864"
    green, green_font = "C6EFCE", "276221"
    amber, amber_font = "FFEB9C", "9C5700"
    red, red_font = "FFC7CE", "9C0006"

    header_fill = PatternFill("solid", start_color=navy)
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    body_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def style_score(cell, value: float) -> None:
        if value >= 90:
            cell.fill = PatternFill("solid", start_color=green)
            cell.font = Font(bold=True, color=green_font, name="Arial", size=10)
        elif value >= 75:
            cell.fill = PatternFill("solid", start_color=amber)
            cell.font = Font(bold=True, color=amber_font, name="Arial", size=10)
        else:
            cell.fill = PatternFill("solid", start_color=red)
            cell.font = Font(bold=True, color=red_font, name="Arial", size=10)

    def write_table(ws, df: pd.DataFrame, percent_cols: set[str] | None = None) -> None:
        percent_cols = percent_cols or set()
        ws.freeze_panes = "A2"
        ws.append(list(df.columns))
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        for _, row in df.iterrows():
            ws.append([
                f"{row[col]:.1f}%" if col in percent_cols and pd.notna(row[col]) else row[col]
                for col in df.columns
            ])
            for cell in ws[ws.max_row]:
                cell.font = body_font
                cell.alignment = left if isinstance(cell.value, str) and len(str(cell.value)) > 20 else center
                cell.border = border
        for column_cells in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 48)

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "CRM Health Score"
    summary = pd.DataFrame([
        {
            "dimension": dimension.capitalize(),
            "weight": f"{WEIGHTS[dimension]:.0%}",
            "overall_score": all_scores.get(f"overall_{dimension}", 100.0),
            "grade": _grade(all_scores.get(f"overall_{dimension}", 100.0)),
        }
        for dimension in WEIGHTS
    ] + [{
        "dimension": "CRM HEALTH SCORE",
        "weight": "100%",
        "overall_score": crm_health,
        "grade": _grade(crm_health),
    }])
    write_table(ws, summary, {"overall_score"})
    for row_num in range(2, ws.max_row + 1):
        raw = ws.cell(row=row_num, column=3).value
        style_score(ws.cell(row=row_num, column=3), float(str(raw).replace("%", "")))
        style_score(ws.cell(row=row_num, column=4), float(str(raw).replace("%", "")))

    ws = wb.create_sheet("Scores by Entity")
    write_table(ws, scores_df, {"score"})

    ws = wb.create_sheet("DQ Issue Log")
    write_table(ws, issue_df.head(10000) if len(issue_df) > 0 else _empty_issues())

    if len(issue_df) > 0:
        check_summary = (
            issue_df.groupby(["dimension", "entity_type", "dq_check", "severity"])
            .size()
            .reset_index(name="issue_count")
            .sort_values("issue_count", ascending=False)
        )
    else:
        check_summary = pd.DataFrame(columns=["dimension", "entity_type", "dq_check", "severity", "issue_count"])
    ws = wb.create_sheet("Check Summary")
    write_table(ws, check_summary)

    for sheet_name, df in record_scores.items():
        ws = wb.create_sheet(sheet_name[:31])
        write_table(ws, df.head(10000), {
            "completeness_score", "validity_score", "consistency_score",
            "uniqueness_score", "freshness_score", "record_quality_score",
        })

    ws = wb.create_sheet("Run Info")
    run_info = pd.DataFrame([
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Module", "Module 5 - Data Quality Engine"),
        ("Platform", "Enterprise CRM Governance & Analytics Platform"),
        ("CRM Health Score", f"{crm_health:.1f}%"),
        ("Total DQ Issues", len(issue_df)),
        ("Freshness Window Days", FRESHNESS_DAYS),
    ], columns=["item", "value"])
    write_table(ws, run_info)

    wb.save(path)
    log.info("  Data quality report -> %s", path)
    return path


def run_data_quality(dedup_data: dict, std_data: dict) -> dict:
    log.info("=" * 70)
    log.info("MODULE 5 - DATA QUALITY ENGINE")
    log.info("=" * 70)

    accounts = dedup_data.get("dedup_accounts", pd.DataFrame()).copy()
    contacts = dedup_data.get("dedup_contacts", pd.DataFrame()).copy()
    leads = dedup_data.get("dedup_leads", pd.DataFrame()).copy()
    stewardship = std_data.get("std_stewardship", pd.DataFrame()).copy()
    account_dedup_map = dedup_data.get("dedup_map_accounts", pd.DataFrame())

    log.info(
        "  Input: %s accounts | %s contacts | %s leads",
        f"{len(accounts):,}", f"{len(contacts):,}", f"{len(leads):,}",
    )

    all_scores = {}
    issue_frames = []

    checks = [
        ("COMPLETENESS", lambda: check_completeness(accounts, contacts, leads)),
        ("VALIDITY", lambda: check_validity(accounts, contacts, leads)),
        ("CONSISTENCY", lambda: check_consistency(accounts, contacts, leads, account_dedup_map)),
        ("UNIQUENESS", lambda: check_uniqueness(accounts, contacts, leads, stewardship)),
        ("FRESHNESS", lambda: check_freshness(accounts, contacts, leads)),
    ]

    for label, fn in checks:
        log.info("  Dimension: %s", label)
        scores, issues = fn()
        all_scores.update(scores)
        issue_frames.append(issues)

    issue_df = _concat_issues(issue_frames)
    crm_health = compute_crm_health_score(all_scores)
    scores_df = build_scores_df(all_scores, crm_health)

    account_quality = build_record_quality_scores(accounts, "ACCOUNT", "account_id", issue_df)
    contact_quality = build_record_quality_scores(contacts, "CONTACT", "contact_id", issue_df)
    lead_quality = build_record_quality_scores(leads, "LEAD", "lead_id", issue_df)

    scores_path = os.path.join(PROCESSED_DIR, "dq_scores.csv")
    issues_path = os.path.join(PROCESSED_DIR, "dq_issue_log.csv")
    account_quality_path = os.path.join(PROCESSED_DIR, "quality_scores_accounts.csv")
    contact_quality_path = os.path.join(PROCESSED_DIR, "quality_scores_contacts.csv")
    lead_quality_path = os.path.join(PROCESSED_DIR, "quality_scores_leads.csv")

    scores_df.to_csv(scores_path, index=False)
    issue_df.to_csv(issues_path, index=False)
    account_quality.to_csv(account_quality_path, index=False)
    contact_quality.to_csv(contact_quality_path, index=False)
    lead_quality.to_csv(lead_quality_path, index=False)

    record_scores = {
        "Account Quality": account_quality,
        "Contact Quality": contact_quality,
        "Lead Quality": lead_quality,
    }
    report_path = write_data_quality_report(scores_df, issue_df, record_scores, crm_health, all_scores)

    log.info("-" * 70)
    for dimension in WEIGHTS:
        log.info(
            "  %-13s %.1f%% (weight %.0f%%)",
            dimension.capitalize(),
            all_scores.get(f"overall_{dimension}", 0.0),
            WEIGHTS[dimension] * 100,
        )
    log.info("  CRM Health    %.1f%%", crm_health)
    log.info("=" * 70)
    log.info("MODULE 5 COMPLETE")
    log.info("=" * 70)

    return {
        "crm_health_score": crm_health,
        "scores_df": scores_df,
        "issue_df": issue_df,
        "all_scores": all_scores,
        "quality_scores_accounts": account_quality,
        "quality_scores_contacts": contact_quality,
        "quality_scores_leads": lead_quality,
        "record_quality_scores": {
            "accounts": account_quality,
            "contacts": contact_quality,
            "leads": lead_quality,
        },
        "output_paths": {
            "dq_scores": scores_path,
            "dq_issue_log": issues_path,
            "quality_scores_accounts": account_quality_path,
            "quality_scores_contacts": contact_quality_path,
            "quality_scores_leads": lead_quality_path,
            "data_quality_report": report_path,
        },
    }


if __name__ == "__main__":
    from module1_ingestion import run_ingestion
    from module2_standardization import run_standardization
    from module3_entity_resolution import run_entity_resolution
    from module4_deduplication import run_deduplication

    staging_data = run_ingestion()
    standardized_data = run_standardization(staging_data)
    entity_match_table = run_entity_resolution(standardized_data)
    deduplicated_data = run_deduplication(standardized_data, entity_match_table)
    run_data_quality(deduplicated_data, standardized_data)
