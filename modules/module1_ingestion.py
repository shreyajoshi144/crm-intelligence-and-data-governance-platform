"""
═══════════════════════════════════════════════════════════════════════════════
MODULE 1 — DATA INGESTION LAYER
Enterprise CRM Governance & Analytics Platform
═══════════════════════════════════════════════════════════════════════════════

PURPOSE:
    Read every source sheet from the master Excel file, validate its schema
    and datatypes, log ingestion metadata, and write clean staging DataFrames
    to CSV files under data/staging/.

OUTPUTS:
    data/staging/stg_accounts.csv
    data/staging/stg_contacts.csv
    data/staging/stg_leads.csv
    data/staging/stg_stewardship.csv
    data/staging/stg_golden_map.csv

    logs/ingestion_log.csv            ← per-sheet load metadata
    reports/schema_validation_report.xlsx  ← column-level type audit
"""

import os
import sys
import logging
import warnings
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings("ignore")

# ── Make the project root importable ──────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config.config import (
    SOURCE_FILE, SOURCE_SHEETS, EXPECTED_SCHEMAS,
    STAGING_DIR, LOGS_DIR, REPORTS_DIR
)

# ── Logger setup ──────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(os.path.join(LOGS_DIR, "module1_ingestion.log"), mode="w"),
    ]
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — Load Excel source file
# ─────────────────────────────────────────────────────────────────────────────

def load_source_file(filepath: str) -> dict[str, pd.DataFrame]:
    """
    Read ALL sheets from the Excel source file.

    Returns:
        dict  {sheet_name: DataFrame}
    """
    log.info("=" * 70)
    log.info("MODULE 1 — DATA INGESTION LAYER")
    log.info("=" * 70)
    log.info(f"Source file : {filepath}")

    if not os.path.exists(filepath):
        log.error(f"Source file NOT found: {filepath}")
        raise FileNotFoundError(f"Source file not found: {filepath}")

    log.info("Loading all sheets from source file...")
    all_sheets = pd.read_excel(filepath, sheet_name=None, dtype=str)
    log.info(f"Sheets found: {list(all_sheets.keys())}")
    return all_sheets


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — Schema Validation
# ─────────────────────────────────────────────────────────────────────────────

def validate_schema(sheet_name: str, df: pd.DataFrame) -> dict:
    """
    Compare actual columns against the expected schema.

    Returns a dict with:
        status          : PASS / FAIL
        missing_columns : columns expected but not present
        extra_columns   : columns present but not expected
        column_details  : per-column type match info
    """
    expected = EXPECTED_SCHEMAS.get(sheet_name, {})
    actual_cols = set(df.columns.str.strip().str.lower())
    expected_cols = set(expected.keys())

    missing = list(expected_cols - actual_cols)
    extra   = list(actual_cols - expected_cols)

    column_details = []
    for col, exp_type in expected.items():
        present = col in df.columns
        if present:
            # Cast numeric columns; flag mismatches
            actual_type = str(df[col].dtype)
            match = "YES" if (actual_type == exp_type or "object" in [actual_type, exp_type]) else "NO"
        else:
            actual_type, match = "MISSING", "NO"
        column_details.append({
            "sheet":         sheet_name,
            "column":        col,
            "expected_type": exp_type,
            "actual_type":   actual_type,
            "type_match":    match,
            "present":       "YES" if present else "NO",
        })

    status = "PASS" if not missing else "FAIL"
    return {
        "status":          status,
        "missing_columns": missing,
        "extra_columns":   extra,
        "column_details":  column_details,
    }


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — Ingest each sheet
# ─────────────────────────────────────────────────────────────────────────────

def ingest_sheet(
    sheet_name: str,
    df: pd.DataFrame,
    staging_name: str,
) -> tuple[pd.DataFrame, dict]:
    """
    Clean column names, cast numeric columns, strip whitespace,
    write staging CSV, and return ingestion metadata.

    Returns:
        (cleaned_df, log_record_dict)
    """
    log.info(f"  ► Ingesting sheet: {sheet_name}  →  {staging_name}")

    # ── Basic column housekeeping ──────────────────────────────────────────
    df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
    df = df.where(pd.notnull(df), None)  # Normalize NaN → None

    # Strip leading/trailing whitespace from all string columns
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].str.strip()

    # Cast numeric columns according to schema
    schema = EXPECTED_SCHEMAS.get(sheet_name, {})
    cast_errors = []
    for col, dtype in schema.items():
        if col in df.columns and dtype in ("float64", "int64"):
            try:
                df[col] = pd.to_numeric(df[col], errors="coerce")
            except Exception as e:
                cast_errors.append(f"{col}: {e}")

    row_count   = len(df)
    null_counts = df.isnull().sum().to_dict()
    total_nulls = sum(null_counts.values())

    # ── Write staging CSV ──────────────────────────────────────────────────
    out_path = os.path.join(STAGING_DIR, f"{staging_name}.csv")
    df.to_csv(out_path, index=False)
    log.info(f"    Rows: {row_count:,}  |  Nulls: {total_nulls}  |  Saved → {out_path}")

    ingestion_record = {
        "timestamp":       datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_sheet":    sheet_name,
        "staging_table":   staging_name,
        "rows_loaded":     row_count,
        "columns":         len(df.columns),
        "total_nulls":     total_nulls,
        "cast_errors":     "; ".join(cast_errors) if cast_errors else "None",
        "status":          "SUCCESS",
        "output_file":     out_path,
    }
    return df, ingestion_record


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — Write Deliverables
# ─────────────────────────────────────────────────────────────────────────────

def write_ingestion_log(log_records: list[dict]) -> str:
    """Save ingestion metadata to CSV."""
    path = os.path.join(LOGS_DIR, "ingestion_log.csv")
    pd.DataFrame(log_records).to_csv(path, index=False)
    log.info(f"  Ingestion log saved → {path}")
    return path


def write_schema_validation_report(all_column_details: list[dict]) -> str:
    """
    Build a formatted Excel schema validation report with:
      - Sheet 1: Summary (one row per sheet)
      - Sheet 2: Column Detail (one row per column per sheet)
    """
    path = os.path.join(REPORTS_DIR, "schema_validation_report.xlsx")

    # ── Summary rows ──────────────────────────────────────────────────────
    sheet_summaries = {}
    for detail in all_column_details:
        sn = detail["sheet"]
        if sn not in sheet_summaries:
            sheet_summaries[sn] = {"total": 0, "pass": 0, "fail": 0}
        sheet_summaries[sn]["total"] += 1
        if detail["type_match"] == "YES" and detail["present"] == "YES":
            sheet_summaries[sn]["pass"] += 1
        else:
            sheet_summaries[sn]["fail"] += 1

    summary_rows = [
        {
            "Sheet":           s,
            "Total Columns":   v["total"],
            "Columns OK":      v["pass"],
            "Columns Failed":  v["fail"],
            "Schema Status":   "PASS" if v["fail"] == 0 else "FAIL",
        }
        for s, v in sheet_summaries.items()
    ]

    # ── openpyxl styles ───────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", start_color="1F3864")  # dark navy
    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    PASS_FILL = PatternFill("solid", start_color="C6EFCE")
    FAIL_FILL = PatternFill("solid", start_color="FFC7CE")
    PASS_FONT = Font(bold=True, color="276221", name="Arial", size=10)
    FAIL_FONT = Font(bold=True, color="9C0006", name="Arial", size=10)
    BODY_FONT = Font(name="Arial", size=10)
    CENTER    = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="CCCCCC")
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.freeze_panes = "A2"

    summary_headers = ["Sheet", "Total Columns", "Columns OK", "Columns Failed", "Schema Status"]
    ws1.append(summary_headers)
    for cell in ws1[1]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for row in summary_rows:
        ws1.append([row[h] for h in summary_headers])
        r = ws1.max_row
        status_cell = ws1.cell(row=r, column=5)
        if status_cell.value == "PASS":
            status_cell.fill = PASS_FILL
            status_cell.font = PASS_FONT
        else:
            status_cell.fill = FAIL_FILL
            status_cell.font = FAIL_FONT
        for c in ws1[r]:
            c.font = BODY_FONT
            c.alignment = CENTER
            c.border = BORDER

    for col in ["A", "B", "C", "D", "E"]:
        ws1.column_dimensions[col].width = 22

    # ── Sheet 2: Column Detail ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Column Detail")
    ws2.freeze_panes = "A2"

    detail_headers = ["Sheet", "Column", "Expected Type", "Actual Type", "Type Match", "Present"]
    ws2.append(detail_headers)
    for cell in ws2[1]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for detail in all_column_details:
        ws2.append([detail[h.lower().replace(" ", "_")] for h in detail_headers])
        r = ws2.max_row
        match_cell = ws2.cell(row=r, column=5)
        if match_cell.value == "YES":
            match_cell.fill = PASS_FILL
            match_cell.font = PASS_FONT
        else:
            match_cell.fill = FAIL_FILL
            match_cell.font = FAIL_FONT
        for c in ws2[r]:
            if c.column != 5:
                c.font = BODY_FONT
            c.alignment = CENTER
            c.border = BORDER

    for col in ["A", "B", "C", "D", "E", "F"]:
        ws2.column_dimensions[col].width = 20

    wb.save(path)
    log.info(f"  Schema validation report saved → {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ORCHESTRATOR
# ─────────────────────────────────────────────────────────────────────────────

def run_ingestion() -> dict[str, pd.DataFrame]:
    """
    End-to-end Module 1 execution.

    Returns:
        dict {staging_name: DataFrame}  ← passed directly into Module 2
    """
    log.info("Starting Module 1 — Data Ingestion")
    os.makedirs(STAGING_DIR, exist_ok=True)
    os.makedirs(LOGS_DIR, exist_ok=True)
    os.makedirs(REPORTS_DIR, exist_ok=True)

    # Step 1 — Load
    all_sheets = load_source_file(SOURCE_FILE)

    ingestion_log      = []
    all_column_details = []
    staging_data       = {}

    for sheet_name, staging_name in SOURCE_SHEETS.items():
        if sheet_name not in all_sheets:
            log.warning(f"  Sheet '{sheet_name}' not found in source file — skipping")
            continue

        df = all_sheets[sheet_name].copy()

        # Step 2 — Validate schema
        validation = validate_schema(sheet_name, df)
        all_column_details.extend(validation["column_details"])
        if validation["missing_columns"]:
            log.warning(f"  ⚠ Missing columns in {sheet_name}: {validation['missing_columns']}")
        else:
            log.info(f"  ✓ Schema OK for {sheet_name}")

        # Step 3 — Ingest
        clean_df, log_record = ingest_sheet(sheet_name, df, staging_name)
        staging_data[staging_name] = clean_df
        ingestion_log.append(log_record)

    # Step 4 — Write deliverables
    log.info("-" * 70)
    log.info("Writing deliverables...")
    write_ingestion_log(ingestion_log)
    write_schema_validation_report(all_column_details)

    log.info("=" * 70)
    log.info(f"MODULE 1 COMPLETE — {len(staging_data)} tables staged successfully")
    log.info("=" * 70)

    return staging_data


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    run_ingestion()
