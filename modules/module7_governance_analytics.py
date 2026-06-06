"""
MODULE 7 - GOVERNANCE ANALYTICS LAYER
Enterprise CRM Governance & Analytics Platform

Purpose
-------
Create governance KPIs for Data Governance teams, CRM administrators,
Sales Operations, and leadership to monitor CRM health.

Outputs:
    data/processed/governance_metrics.csv
    reports/governance_report.xlsx
    logs/module7_governance.log

Public entry point:
    run_governance_analytics(dq_data=None, golden_data=None, dedup_data=None, std_data=None) -> dict
"""

from __future__ import annotations

import logging
import os
import sys
import warnings
from datetime import date, datetime
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

warnings.filterwarnings("ignore")

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from config.config import LOGS_DIR, PROCESSED_DIR, REPORTS_DIR
except Exception:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    PROCESSED_DIR = os.path.join(BASE_DIR, "data", "processed")
    LOGS_DIR = os.path.join(BASE_DIR, "logs")
    REPORTS_DIR = os.path.join(BASE_DIR, "reports")


for directory in (PROCESSED_DIR, LOGS_DIR, REPORTS_DIR):
    os.makedirs(directory, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(os.path.join(LOGS_DIR, "module7_governance.log"), mode="w"),
    ],
)
log = logging.getLogger(__name__)


REPORTING_DATE = date.today().isoformat()


ENTITY_CONFIG = {
    "accounts": {
        "label": "Accounts",
        "dedup_key": "dedup_accounts",
        "dedup_file": "dedup_accounts.csv",
        "dedup_map_key": "dedup_map_accounts",
        "dedup_map_file": "dedup_map_accounts.csv",
        "quality_key": "quality_scores_accounts",
        "quality_file": "quality_scores_accounts.csv",
        "golden_key": "golden_accounts",
        "golden_file": "golden_accounts.csv",
        "golden_source_ids_col": "source_account_ids",
    },
    "contacts": {
        "label": "Contacts",
        "dedup_key": "dedup_contacts",
        "dedup_file": "dedup_contacts.csv",
        "dedup_map_key": "dedup_map_contacts",
        "dedup_map_file": "dedup_map_contacts.csv",
        "quality_key": "quality_scores_contacts",
        "quality_file": "quality_scores_contacts.csv",
        "golden_key": "golden_contacts",
        "golden_file": "golden_contacts.csv",
        "golden_source_ids_col": "source_contact_ids",
    },
    "leads": {
        "label": "Leads",
        "dedup_key": "dedup_leads",
        "dedup_file": "dedup_leads.csv",
        "dedup_map_key": "dedup_map_leads",
        "dedup_map_file": "dedup_map_leads.csv",
        "quality_key": "quality_scores_leads",
        "quality_file": "quality_scores_leads.csv",
        "golden_key": "golden_leads",
        "golden_file": "golden_leads.csv",
        "golden_source_ids_col": "source_lead_ids",
    },
}


DIMENSIONS = ["Completeness", "Validity", "Consistency", "Uniqueness", "Freshness"]


def _csv_path(filename: str) -> str:
    return os.path.join(PROCESSED_DIR, filename)


def _read_processed_csv(filename: str) -> pd.DataFrame:
    path = _csv_path(filename)
    if not os.path.exists(path):
        log.warning("  Missing optional input file: %s", path)
        return pd.DataFrame()
    df = pd.read_csv(path)
    log.info("  Loaded %-34s %8s rows", filename, f"{len(df):,}")
    return df


def _frame_from_sources(
    primary: dict[str, Any] | None,
    secondary: dict[str, Any] | None,
    key: str,
    filename: str,
) -> pd.DataFrame:
    for source in (primary, secondary):
        if isinstance(source, dict) and key in source and source[key] is not None:
            return source[key].copy()
    return _read_processed_csv(filename)


def _safe_numeric(value: Any, default: float = 0.0) -> float:
    parsed = pd.to_numeric(value, errors="coerce")
    if pd.isna(parsed):
        return default
    return float(parsed)


def _metric(metric_name: str, metric_value: Any, metric_category: str) -> dict[str, Any]:
    return {
        "metric_name": metric_name,
        "metric_value": metric_value,
        "metric_category": metric_category,
        "reporting_date": REPORTING_DATE,
    }


def _score_from_dq_scores(dq_scores: pd.DataFrame, metric_name: str) -> float:
    if len(dq_scores) == 0:
        return 0.0

    entity = dq_scores.get("entity", pd.Series(dtype=str)).astype(str).str.upper()
    dimension = dq_scores.get("dimension", pd.Series(dtype=str)).astype(str).str.lower()

    if metric_name == "CRM Health Score":
        mask = entity.eq("CRM HEALTH SCORE")
        if mask.any() and "score" in dq_scores.columns:
            return round(_safe_numeric(dq_scores.loc[mask, "score"].iloc[0]), 2)
        mask = dimension.eq("weighted average")
        if mask.any() and "score" in dq_scores.columns:
            return round(_safe_numeric(dq_scores.loc[mask, "score"].iloc[0]), 2)
        return 0.0

    target_dimension = metric_name.replace("Data ", "").replace(" %", "").lower()
    mask = entity.isin(["ACCOUNTS", "CONTACTS", "LEADS"]) & dimension.eq(target_dimension)
    if mask.any() and "score" in dq_scores.columns:
        return round(pd.to_numeric(dq_scores.loc[mask, "score"], errors="coerce").mean(), 2)

    mask = entity.eq("OVERALL") & dimension.eq(target_dimension)
    if mask.any() and "score" in dq_scores.columns:
        return round(_safe_numeric(dq_scores.loc[mask, "score"].iloc[0]), 2)

    return 0.0


def _dedup_counts(dedup_map: pd.DataFrame, dedup_df: pd.DataFrame) -> tuple[int, int, int]:
    if len(dedup_map) > 0 and {"record_id", "canonical_id"}.issubset(dedup_map.columns):
        total_before = int(dedup_map["record_id"].dropna().astype(str).nunique())
        total_after = int(dedup_map["canonical_id"].dropna().astype(str).nunique())
        records_merged = max(total_before - total_after, 0)
        return total_before, total_after, records_merged

    total_after = len(dedup_df)
    return total_after, total_after, 0


def _source_id_count_from_golden(golden_df: pd.DataFrame, source_ids_col: str) -> int:
    if len(golden_df) == 0 or source_ids_col not in golden_df.columns:
        return 0
    ids: set[str] = set()
    for value in golden_df[source_ids_col].dropna():
        for record_id in str(value).split(","):
            record_id = record_id.strip()
            if record_id:
                ids.add(record_id)
    return len(ids)


def _golden_coverage(golden_df: pd.DataFrame, dedup_map: pd.DataFrame, dedup_df: pd.DataFrame, source_ids_col: str) -> tuple[int, int, float]:
    total_records = _source_id_count_from_golden(golden_df, source_ids_col)
    if total_records == 0:
        total_records, _, _ = _dedup_counts(dedup_map, dedup_df)

    golden_records = len(golden_df)
    coverage = round((golden_records / total_records) * 100, 2) if total_records else 0.0
    return total_records, golden_records, coverage


def _quality_score_series(quality_frames: dict[str, pd.DataFrame]) -> pd.Series:
    series_list = []
    for df in quality_frames.values():
        if len(df) > 0 and "record_quality_score" in df.columns:
            series_list.append(pd.to_numeric(df["record_quality_score"], errors="coerce"))
    if not series_list:
        return pd.Series(dtype=float)
    return pd.concat(series_list, ignore_index=True).dropna()


def _quality_band_distribution(scores: pd.Series) -> pd.DataFrame:
    bands = [
        ("Excellent", scores >= 90),
        ("Good", (scores >= 75) & (scores < 90)),
        ("Fair", (scores >= 60) & (scores < 75)),
        ("Poor", scores < 60),
    ]
    total = len(scores)
    rows = []
    for band, mask in bands:
        count = int(mask.sum()) if total else 0
        pct = round((count / total) * 100, 2) if total else 0.0
        rows.append({"quality_band": band, "record_count": count, "percentage": pct})
    return pd.DataFrame(rows)


def _stewardship_count(std_stewardship: pd.DataFrame) -> int:
    if len(std_stewardship) == 0 or "dq_issue_status" not in std_stewardship.columns:
        return 0
    return int(std_stewardship["dq_issue_status"].isin(["Open", "In Review", "Escalated"]).sum())


def build_governance_metrics(
    dq_scores: pd.DataFrame,
    quality_frames: dict[str, pd.DataFrame],
    golden_frames: dict[str, pd.DataFrame],
    dedup_frames: dict[str, pd.DataFrame],
    dedup_map_frames: dict[str, pd.DataFrame],
    std_stewardship: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    metrics: list[dict[str, Any]] = []

    metrics.append(_metric("CRM Health Score", _score_from_dq_scores(dq_scores, "CRM Health Score"), "Overall"))
    for dimension in DIMENSIONS:
        metrics.append(_metric(
            f"Data {dimension} %",
            _score_from_dq_scores(dq_scores, f"Data {dimension} %"),
            "Data Quality",
        ))

    total_before = 0
    total_after = 0
    total_merged = 0
    golden_summary_rows = []

    for entity_key, cfg in ENTITY_CONFIG.items():
        dedup_df = dedup_frames[entity_key]
        dedup_map = dedup_map_frames[entity_key]
        before, after, merged = _dedup_counts(dedup_map, dedup_df)
        total_before += before
        total_after += after
        total_merged += merged

        total_records, golden_records, coverage = _golden_coverage(
            golden_frames[entity_key], dedup_map, dedup_df, cfg["golden_source_ids_col"],
        )
        golden_summary_rows.append({
            "entity_type": cfg["label"],
            "total_records": total_records,
            "golden_records": golden_records,
            "coverage_pct": coverage,
        })
        metrics.append(_metric(f"Golden Record Coverage % - {cfg['label']}", coverage, "Golden Records"))

    duplicate_rate = round((total_merged / total_before) * 100, 2) if total_before else 0.0
    metrics.append(_metric("Duplicate Rate", duplicate_rate, "Deduplication"))

    overall_total_records = sum(row["total_records"] for row in golden_summary_rows)
    overall_golden_records = sum(row["golden_records"] for row in golden_summary_rows)
    overall_coverage = round((overall_golden_records / overall_total_records) * 100, 2) if overall_total_records else 0.0
    metrics.append(_metric("Golden Record Coverage % - Overall", overall_coverage, "Golden Records"))

    stewardship_records = _stewardship_count(std_stewardship)
    metrics.append(_metric("Records Under Stewardship", stewardship_records, "Governance Issues"))

    scores = _quality_score_series(quality_frames)
    requiring_review = int((scores < 75).sum()) if len(scores) else 0
    excellent_records = int((scores >= 90).sum()) if len(scores) else 0
    poor_records = int((scores < 60).sum()) if len(scores) else 0

    metrics.append(_metric("Records Requiring Review", requiring_review, "Governance Issues"))
    metrics.append(_metric("Excellent Records", excellent_records, "Data Quality"))
    metrics.append(_metric("Poor Quality Records", poor_records, "Governance Issues"))

    metrics_df = pd.DataFrame(metrics, columns=[
        "metric_name", "metric_value", "metric_category", "reporting_date",
    ])

    quality_distribution = _quality_band_distribution(scores)
    golden_summary = pd.DataFrame(golden_summary_rows)
    golden_summary.loc[len(golden_summary)] = {
        "entity_type": "Overall",
        "total_records": overall_total_records,
        "golden_records": overall_golden_records,
        "coverage_pct": overall_coverage,
    }

    return metrics_df, quality_distribution, golden_summary


def _metric_lookup(metrics_df: pd.DataFrame, metric_name: str) -> Any:
    match = metrics_df.loc[metrics_df["metric_name"].eq(metric_name), "metric_value"]
    if len(match) == 0:
        return 0
    return match.iloc[0]


def write_governance_report(
    metrics_df: pd.DataFrame,
    quality_distribution: pd.DataFrame,
    golden_summary: pd.DataFrame,
    record_counts: dict[str, int],
) -> str:
    path = os.path.join(REPORTS_DIR, "governance_report.xlsx")

    header_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    body_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def write_df(ws, df: pd.DataFrame) -> None:
        ws.freeze_panes = "A2"
        ws.append(list(df.columns))
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        for _, row in df.iterrows():
            ws.append([row[col] for col in df.columns])
            for cell in ws[ws.max_row]:
                cell.font = body_font
                cell.alignment = left if isinstance(cell.value, str) and len(str(cell.value)) > 25 else center
                cell.border = border

        for column_cells in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 52)

    kpi_summary = pd.DataFrame([
        {"kpi": "CRM Health Score", "value": _metric_lookup(metrics_df, "CRM Health Score")},
        {"kpi": "Duplicate Rate", "value": _metric_lookup(metrics_df, "Duplicate Rate")},
        {"kpi": "Golden Record Coverage", "value": _metric_lookup(metrics_df, "Golden Record Coverage % - Overall")},
        {"kpi": "Completeness", "value": _metric_lookup(metrics_df, "Data Completeness %")},
        {"kpi": "Validity", "value": _metric_lookup(metrics_df, "Data Validity %")},
        {"kpi": "Consistency", "value": _metric_lookup(metrics_df, "Data Consistency %")},
        {"kpi": "Uniqueness", "value": _metric_lookup(metrics_df, "Data Uniqueness %")},
        {"kpi": "Freshness", "value": _metric_lookup(metrics_df, "Data Freshness %")},
    ])

    governance_issues = pd.DataFrame([
        {"issue": "Records Requiring Review", "count": _metric_lookup(metrics_df, "Records Requiring Review")},
        {"issue": "Poor Quality Records", "count": _metric_lookup(metrics_df, "Poor Quality Records")},
        {"issue": "Open Stewardship Issues", "count": _metric_lookup(metrics_df, "Records Under Stewardship")},
    ])

    run_information = pd.DataFrame([
        {"item": "Timestamp", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"item": "Module Name", "value": "Module 7 - Governance Analytics"},
        {"item": "Reporting Date", "value": REPORTING_DATE},
        {"item": "Deduplicated Account Records", "value": record_counts.get("dedup_accounts", 0)},
        {"item": "Deduplicated Contact Records", "value": record_counts.get("dedup_contacts", 0)},
        {"item": "Deduplicated Lead Records", "value": record_counts.get("dedup_leads", 0)},
        {"item": "Quality Score Records", "value": record_counts.get("quality_records", 0)},
        {"item": "Governance Metrics", "value": len(metrics_df)},
    ])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Governance KPI Summary"
    write_df(ws, kpi_summary)

    ws = wb.create_sheet("Data Quality Summary")
    write_df(ws, quality_distribution)

    ws = wb.create_sheet("Governance Issues")
    write_df(ws, governance_issues)

    ws = wb.create_sheet("Golden Record Summary")
    write_df(ws, golden_summary)

    ws = wb.create_sheet("Run Information")
    write_df(ws, run_information)

    wb.save(path)
    log.info("  Governance report -> %s", path)
    return path


def run_governance_analytics(
    dq_data: dict[str, Any] | None = None,
    golden_data: dict[str, pd.DataFrame] | None = None,
    dedup_data: dict[str, pd.DataFrame] | None = None,
    std_data: dict[str, pd.DataFrame] | None = None,
) -> dict[str, Any]:
    log.info("=" * 70)
    log.info("MODULE 7 - GOVERNANCE ANALYTICS")
    log.info("=" * 70)

    dq_data = dq_data or {}
    golden_data = golden_data or {}
    dedup_data = dedup_data or {}
    std_data = std_data or {}

    dq_scores = _frame_from_sources(dq_data, None, "scores_df", "dq_scores.csv")
    if len(dq_scores) == 0:
        dq_scores = _frame_from_sources(dq_data, None, "dq_scores", "dq_scores.csv")

    quality_frames: dict[str, pd.DataFrame] = {}
    golden_frames: dict[str, pd.DataFrame] = {}
    dedup_frames: dict[str, pd.DataFrame] = {}
    dedup_map_frames: dict[str, pd.DataFrame] = {}

    for entity_key, cfg in ENTITY_CONFIG.items():
        quality_frames[entity_key] = _frame_from_sources(dq_data, None, cfg["quality_key"], cfg["quality_file"])
        golden_frames[entity_key] = _frame_from_sources(golden_data, None, cfg["golden_key"], cfg["golden_file"])
        dedup_frames[entity_key] = _frame_from_sources(dedup_data, None, cfg["dedup_key"], cfg["dedup_file"])
        dedup_map_frames[entity_key] = _frame_from_sources(dedup_data, None, cfg["dedup_map_key"], cfg["dedup_map_file"])

    std_stewardship = _frame_from_sources(std_data, None, "std_stewardship", "std_stewardship.csv")

    metrics_df, quality_distribution, golden_summary = build_governance_metrics(
        dq_scores=dq_scores,
        quality_frames=quality_frames,
        golden_frames=golden_frames,
        dedup_frames=dedup_frames,
        dedup_map_frames=dedup_map_frames,
        std_stewardship=std_stewardship,
    )

    metrics_path = os.path.join(PROCESSED_DIR, "governance_metrics.csv")
    metrics_df.to_csv(metrics_path, index=False)
    log.info("  governance_metrics.csv -> %s (%s rows)", metrics_path, f"{len(metrics_df):,}")

    record_counts = {
        "dedup_accounts": len(dedup_frames["accounts"]),
        "dedup_contacts": len(dedup_frames["contacts"]),
        "dedup_leads": len(dedup_frames["leads"]),
        "quality_records": sum(len(df) for df in quality_frames.values()),
    }
    report_path = write_governance_report(metrics_df, quality_distribution, golden_summary, record_counts)

    log.info("-" * 70)
    log.info("GOVERNANCE KPI SUMMARY")
    for metric_name in [
        "CRM Health Score", "Duplicate Rate", "Golden Record Coverage % - Overall",
        "Data Completeness %", "Data Validity %", "Data Consistency %",
        "Data Uniqueness %", "Data Freshness %",
    ]:
        log.info("  %-34s %s", metric_name, _metric_lookup(metrics_df, metric_name))
    log.info("=" * 70)
    log.info("MODULE 7 COMPLETE")
    log.info("=" * 70)

    return {
        "governance_metrics": metrics_df,
        "governance_report_path": report_path,
        "output_paths": {
            "governance_metrics": metrics_path,
            "governance_report": report_path,
            "module7_log": os.path.join(LOGS_DIR, "module7_governance.log"),
        },
    }


if __name__ == "__main__":
    run_governance_analytics()
