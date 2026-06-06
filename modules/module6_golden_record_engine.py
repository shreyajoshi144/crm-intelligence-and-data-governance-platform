"""
MODULE 6 - GOLDEN RECORD ENGINE
Enterprise CRM Governance & Analytics Platform

Purpose
-------
Create master CRM Golden Records using:
    - deduplicated datasets
    - deduplication cluster maps
    - Module 5 record-level quality scores
    - standardised golden record mappings

Outputs:
    data/processed/golden_accounts.csv
    data/processed/golden_contacts.csv
    data/processed/golden_leads.csv
    reports/golden_record_report.xlsx
    logs/module6_golden_record.log

Public entry point:
    run_golden_record_engine(dedup_data=None, dq_data=None, std_data=None) -> dict
"""

from __future__ import annotations

import logging
import os
import sys
import warnings
from datetime import datetime
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

warnings.filterwarnings("ignore")

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from config.config import LOGS_DIR, PROCESSED_DIR, REPORTS_DIR
except Exception:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    PROCESSED_DIR = os.path.join(BASE_DIR, "data", "processed")
    LOGS_DIR = os.path.join(BASE_DIR, "logs")
    REPORTS_DIR = os.path.join(BASE_DIR, "reports")


for directory in (PROCESSED_DIR, LOGS_DIR, REPORTS_DIR):
    os.makedirs(directory, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(os.path.join(LOGS_DIR, "module6_golden_record.log"), mode="w"),
    ],
)
log = logging.getLogger(__name__)


DATE_COLUMNS = ["last_modified_date", "modified_date", "created_date"]

ENTITY_CONFIG = {
    "accounts": {
        "entity_label": "Accounts",
        "id_col": "account_id",
        "golden_id_col": "golden_account_id",
        "surviving_id_col": "surviving_account_id",
        "source_ids_col": "source_account_ids",
        "golden_prefix": "GOLD_ACC",
        "dedup_key": "dedup_accounts",
        "dedup_map_key": "dedup_map_accounts",
        "quality_key": "quality_scores_accounts",
        "dedup_file": "dedup_accounts.csv",
        "dedup_map_file": "dedup_map_accounts.csv",
        "quality_file": "quality_scores_accounts.csv",
        "output_file": "golden_accounts.csv",
        "output_columns": [
            "golden_account_id", "surviving_account_id", "source_account_ids",
            "account_name", "industry", "country", "annual_revenue", "employee_count",
            "record_quality_score", "quality_grade", "survivorship_reason",
            "records_merged_count",
        ],
        "business_columns": ["account_name", "industry", "country", "annual_revenue", "employee_count"],
    },
    "contacts": {
        "entity_label": "Contacts",
        "id_col": "contact_id",
        "golden_id_col": "golden_contact_id",
        "surviving_id_col": "surviving_contact_id",
        "source_ids_col": "source_contact_ids",
        "golden_prefix": "GOLD_CON",
        "dedup_key": "dedup_contacts",
        "dedup_map_key": "dedup_map_contacts",
        "quality_key": "quality_scores_contacts",
        "dedup_file": "dedup_contacts.csv",
        "dedup_map_file": "dedup_map_contacts.csv",
        "quality_file": "quality_scores_contacts.csv",
        "output_file": "golden_contacts.csv",
        "output_columns": [
            "golden_contact_id", "surviving_contact_id", "source_contact_ids",
            "full_name", "email", "phone", "job_title",
            "record_quality_score", "quality_grade", "survivorship_reason",
            "records_merged_count",
        ],
        "business_columns": ["full_name", "email", "phone", "job_title"],
    },
    "leads": {
        "entity_label": "Leads",
        "id_col": "lead_id",
        "golden_id_col": "golden_lead_id",
        "surviving_id_col": "surviving_lead_id",
        "source_ids_col": "source_lead_ids",
        "golden_prefix": "GOLD_LED",
        "dedup_key": "dedup_leads",
        "dedup_map_key": "dedup_map_leads",
        "quality_key": "quality_scores_leads",
        "dedup_file": "dedup_leads.csv",
        "dedup_map_file": "dedup_map_leads.csv",
        "quality_file": "quality_scores_leads.csv",
        "output_file": "golden_leads.csv",
        "output_columns": [
            "golden_lead_id", "surviving_lead_id", "source_lead_ids",
            "lead_name", "source", "status",
            "record_quality_score", "quality_grade", "survivorship_reason",
            "records_merged_count",
        ],
        "business_columns": ["lead_name", "source", "status"],
    },
}


def _csv_path(filename: str) -> str:
    return os.path.join(PROCESSED_DIR, filename)


def _read_processed_csv(filename: str) -> pd.DataFrame:
    path = _csv_path(filename)
    if not os.path.exists(path):
        log.warning("  Missing optional input file: %s", path)
        return pd.DataFrame()
    df = pd.read_csv(path)
    log.info("  Loaded %-32s %8s rows", filename, f"{len(df):,}")
    return df


def _frame_from_sources(
    primary: dict[str, Any] | None,
    secondary: dict[str, Any] | None,
    key: str,
    filename: str,
) -> pd.DataFrame:
    for source in (primary, secondary):
        if isinstance(source, dict) and key in source and source[key] is not None:
            return source[key].copy()
    return _read_processed_csv(filename)


def _clean_id(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _sort_ids(ids: list[str]) -> list[str]:
    return sorted({_clean_id(value) for value in ids if _clean_id(value)})


def _safe_numeric(value: Any, default: float = -1.0) -> float:
    numeric = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric):
        return default
    return float(numeric)


def _safe_int(value: Any, default: int = 10**9) -> int:
    numeric = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric):
        return default
    return int(numeric)


def _latest_date_value(row: dict[str, Any]) -> pd.Timestamp:
    values = []
    for col in DATE_COLUMNS:
        if col in row:
            parsed = pd.to_datetime(row.get(col), errors="coerce")
            if pd.notna(parsed):
                values.append(parsed)
    if not values:
        return pd.Timestamp.min
    return max(values)


def _latest_date_rank(row: dict[str, Any]) -> int:
    latest = _latest_date_value(row)
    if latest is pd.Timestamp.min or latest == pd.Timestamp.min:
        return 0
    return int(latest.timestamp())


def _quality_grade(score: float) -> str:
    if score >= 90:
        return "A"
    if score >= 80:
        return "B"
    if score >= 70:
        return "C"
    return "D"


def _ensure_quality_columns(quality_df: pd.DataFrame, id_col: str) -> pd.DataFrame:
    out = quality_df.copy()
    if len(out) == 0:
        out = pd.DataFrame(columns=[
            id_col, "record_quality_score", "completeness_score",
            "issue_count", "quality_grade",
        ])
    if id_col not in out.columns:
        for possible in ("record_id", "account_id", "contact_id", "lead_id"):
            if possible in out.columns:
                out = out.rename(columns={possible: id_col})
                break
    if id_col not in out.columns:
        out[id_col] = []

    out[id_col] = out[id_col].astype(str).str.strip()
    defaults = {
        "record_quality_score": 0.0,
        "completeness_score": 0.0,
        "issue_count": 10**9,
        "quality_grade": "D",
    }
    for col, default in defaults.items():
        if col not in out.columns:
            out[col] = default
    return out


def _build_record_pool(master_df: pd.DataFrame, quality_df: pd.DataFrame, id_col: str) -> pd.DataFrame:
    master = master_df.copy()
    quality = _ensure_quality_columns(quality_df, id_col)

    if len(master) == 0 and len(quality) == 0:
        return pd.DataFrame(columns=[id_col])

    if id_col not in master.columns:
        master[id_col] = []
    master[id_col] = master[id_col].astype(str).str.strip()

    pool = master.merge(quality, on=id_col, how="outer", suffixes=("", "_quality"))

    if "quality_grade_quality" in pool.columns and "quality_grade" in pool.columns:
        pool["quality_grade"] = pool["quality_grade"].fillna(pool["quality_grade_quality"])
        pool = pool.drop(columns=["quality_grade_quality"])

    for col in ("record_quality_score", "completeness_score", "issue_count"):
        pool[col] = pd.to_numeric(pool[col], errors="coerce")

    pool["record_quality_score"] = pool["record_quality_score"].fillna(0.0)
    pool["completeness_score"] = pool["completeness_score"].fillna(0.0)
    pool["issue_count"] = pool["issue_count"].fillna(10**9).astype(int)
    pool["quality_grade"] = pool["quality_grade"].fillna(pool["record_quality_score"].apply(_quality_grade))

    return pool


def _build_clusters(master_df: pd.DataFrame, map_df: pd.DataFrame, id_col: str) -> list[dict[str, Any]]:
    clusters: list[dict[str, Any]] = []
    seen_ids: set[str] = set()

    if len(map_df) > 0 and {"record_id", "canonical_id"}.issubset(map_df.columns):
        cleaned = map_df.copy()
        cleaned["record_id"] = cleaned["record_id"].astype(str).str.strip()
        cleaned["canonical_id"] = cleaned["canonical_id"].astype(str).str.strip()

        for canonical_id, group in cleaned.groupby("canonical_id", dropna=False):
            source_ids = _sort_ids(group["record_id"].tolist() + [canonical_id])
            if not source_ids:
                continue
            clusters.append({
                "canonical_id": _clean_id(canonical_id) or source_ids[0],
                "source_ids": source_ids,
            })
            seen_ids.update(source_ids)

    if id_col in master_df.columns:
        for record_id in master_df[id_col].dropna().astype(str).str.strip():
            if record_id and record_id not in seen_ids:
                clusters.append({
                    "canonical_id": record_id,
                    "source_ids": [record_id],
                })
                seen_ids.add(record_id)

    return clusters


def _golden_map_lookup(std_golden_map: pd.DataFrame, id_col: str) -> tuple[dict[str, str], dict[str, list[str]]]:
    if std_golden_map is None or len(std_golden_map) == 0:
        return {}, {}

    candidates = [
        ("account_id", "golden_record_id"),
        ("contact_id", "golden_contact_id"),
        ("lead_id", "golden_lead_id"),
        (id_col, "golden_record_id"),
        (id_col, f"golden_{id_col}"),
    ]

    record_col = None
    golden_col = None
    for possible_record_col, possible_golden_col in candidates:
        if possible_record_col in std_golden_map.columns and possible_golden_col in std_golden_map.columns:
            record_col = possible_record_col
            golden_col = possible_golden_col
            break

    if record_col is None or golden_col is None:
        return {}, {}

    mapping: dict[str, str] = {}
    reverse: dict[str, list[str]] = {}
    for _, row in std_golden_map.iterrows():
        record_id = _clean_id(row.get(record_col))
        golden_id = _clean_id(row.get(golden_col))
        if not record_id or not golden_id:
            continue
        mapping[record_id] = golden_id
        reverse.setdefault(golden_id, []).append(record_id)

    return mapping, reverse


def _select_survivor(
    cluster_ids: list[str],
    canonical_id: str,
    pool_by_id: dict[str, dict[str, Any]],
    golden_lookup: dict[str, str],
) -> tuple[str, str, str | None]:
    available_ids = [record_id for record_id in cluster_ids if record_id in pool_by_id]
    if not available_ids and canonical_id:
        available_ids = [canonical_id]
    if not available_ids:
        available_ids = cluster_ids[:1]

    mapped_ids = [record_id for record_id in available_ids if record_id in golden_lookup]
    if mapped_ids:
        survivor = sorted(mapped_ids)[0]
        return survivor, "Golden Record Map", golden_lookup.get(survivor)

    scored = []
    for record_id in available_ids:
        row = pool_by_id.get(record_id, {})
        scored.append({
            "record_id": record_id,
            "record_quality_score": _safe_numeric(row.get("record_quality_score"), -1.0),
            "completeness_score": _safe_numeric(row.get("completeness_score"), -1.0),
            "issue_count": _safe_int(row.get("issue_count"), 10**9),
            "latest_date_rank": _latest_date_rank(row),
        })

    if not scored:
        fallback = _sort_ids(cluster_ids)[0]
        return fallback, "Lowest ID", None

    max_quality = max(row["record_quality_score"] for row in scored)
    candidates = [row for row in scored if row["record_quality_score"] == max_quality]
    if len(candidates) == 1:
        return candidates[0]["record_id"], "Highest record_quality_score", None

    max_completeness = max(row["completeness_score"] for row in candidates)
    candidates = [row for row in candidates if row["completeness_score"] == max_completeness]
    if len(candidates) == 1:
        return candidates[0]["record_id"], "Highest completeness_score", None

    min_issues = min(row["issue_count"] for row in candidates)
    candidates = [row for row in candidates if row["issue_count"] == min_issues]
    if len(candidates) == 1:
        return candidates[0]["record_id"], "Fewest DQ issues", None

    max_date = max(row["latest_date_rank"] for row in candidates)
    candidates = [row for row in candidates if row["latest_date_rank"] == max_date]
    if len(candidates) == 1 and max_date > 0:
        return candidates[0]["record_id"], "Most recent record", None

    survivor = sorted(row["record_id"] for row in candidates)[0]
    return survivor, "Lowest ID", None


def _build_golden_records_for_entity(
    entity_key: str,
    master_df: pd.DataFrame,
    map_df: pd.DataFrame,
    quality_df: pd.DataFrame,
    std_golden_map: pd.DataFrame,
) -> pd.DataFrame:
    cfg = ENTITY_CONFIG[entity_key]
    id_col = cfg["id_col"]

    pool = _build_record_pool(master_df, quality_df, id_col)
    pool_by_id = {
        _clean_id(row[id_col]): row.to_dict()
        for _, row in pool.iterrows()
        if _clean_id(row.get(id_col))
    }

    clusters = _build_clusters(master_df, map_df, id_col)
    golden_lookup, _ = _golden_map_lookup(std_golden_map, id_col)

    rows = []
    for idx, cluster in enumerate(clusters, start=1):
        source_ids = _sort_ids(cluster["source_ids"])
        canonical_id = _clean_id(cluster.get("canonical_id")) or (source_ids[0] if source_ids else "")
        survivor_id, reason, mapped_golden_id = _select_survivor(
            source_ids, canonical_id, pool_by_id, golden_lookup,
        )

        survivor_row = pool_by_id.get(survivor_id) or pool_by_id.get(canonical_id) or {}
        generated_golden_id = f"{cfg['golden_prefix']}_{idx:06d}"
        golden_id = mapped_golden_id or generated_golden_id

        record = {
            cfg["golden_id_col"]: golden_id,
            cfg["surviving_id_col"]: survivor_id,
            cfg["source_ids_col"]: ",".join(source_ids),
            "record_quality_score": round(_safe_numeric(survivor_row.get("record_quality_score"), 0.0), 2),
            "quality_grade": survivor_row.get("quality_grade") or _quality_grade(_safe_numeric(survivor_row.get("record_quality_score"), 0.0)),
            "survivorship_reason": reason,
            "records_merged_count": len(source_ids),
        }

        for col in cfg["business_columns"]:
            record[col] = survivor_row.get(col, "")

        rows.append(record)

    output = pd.DataFrame(rows)
    for col in cfg["output_columns"]:
        if col not in output.columns:
            output[col] = ""
    return output[cfg["output_columns"]]


def _summary_kpis(
    accounts_df: pd.DataFrame,
    contacts_df: pd.DataFrame,
    leads_df: pd.DataFrame,
    golden_accounts: pd.DataFrame,
    golden_contacts: pd.DataFrame,
    golden_leads: pd.DataFrame,
) -> pd.DataFrame:
    total_accounts = int(golden_accounts.get("records_merged_count", pd.Series(dtype=int)).sum())
    total_contacts = int(golden_contacts.get("records_merged_count", pd.Series(dtype=int)).sum())
    total_leads = int(golden_leads.get("records_merged_count", pd.Series(dtype=int)).sum())
    if total_accounts == 0:
        total_accounts = len(accounts_df)
    if total_contacts == 0:
        total_contacts = len(contacts_df)
    if total_leads == 0:
        total_leads = len(leads_df)

    total_records = total_accounts + total_contacts + total_leads
    total_golden = len(golden_accounts) + len(golden_contacts) + len(golden_leads)
    records_merged = int(total_records - total_golden)
    merge_rate = records_merged / total_records if total_records else 0.0

    quality_values = pd.concat([
        golden_accounts.get("record_quality_score", pd.Series(dtype=float)),
        golden_contacts.get("record_quality_score", pd.Series(dtype=float)),
        golden_leads.get("record_quality_score", pd.Series(dtype=float)),
    ], ignore_index=True)
    avg_quality = round(pd.to_numeric(quality_values, errors="coerce").mean(), 2) if len(quality_values) else 0.0

    return pd.DataFrame([
        {"KPI": "Total Accounts", "Value": total_accounts},
        {"KPI": "Golden Accounts", "Value": len(golden_accounts)},
        {"KPI": "Total Contacts", "Value": total_contacts},
        {"KPI": "Golden Contacts", "Value": len(golden_contacts)},
        {"KPI": "Total Leads", "Value": total_leads},
        {"KPI": "Golden Leads", "Value": len(golden_leads)},
        {"KPI": "Records Merged", "Value": records_merged},
        {"KPI": "Merge Rate", "Value": f"{merge_rate:.2%}"},
        {"KPI": "Average Quality Score", "Value": f"{avg_quality:.2f}%"},
    ])


def write_golden_record_report(
    summary_df: pd.DataFrame,
    golden_accounts: pd.DataFrame,
    golden_contacts: pd.DataFrame,
    golden_leads: pd.DataFrame,
) -> str:
    path = os.path.join(REPORTS_DIR, "golden_record_report.xlsx")

    header_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    body_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def write_df(ws, df: pd.DataFrame) -> None:
        ws.freeze_panes = "A2"
        ws.append(list(df.columns))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for _, row in df.iterrows():
            ws.append([row[col] for col in df.columns])
            for cell in ws[ws.max_row]:
                cell.font = body_font
                cell.alignment = left if isinstance(cell.value, str) and len(str(cell.value)) > 25 else center
                cell.border = border

        for column_cells in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 52)

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Golden Record Summary"
    write_df(ws, summary_df)

    ws = wb.create_sheet("Golden Accounts")
    write_df(ws, golden_accounts)

    ws = wb.create_sheet("Golden Contacts")
    write_df(ws, golden_contacts)

    ws = wb.create_sheet("Golden Leads")
    write_df(ws, golden_leads)

    ws = wb.create_sheet("Run Info")
    run_info = pd.DataFrame([
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Module", "Module 6 - Golden Record Engine"),
        ("Platform", "Enterprise CRM Governance & Analytics Platform"),
        ("Survivorship Priority 1", "Golden Record Map"),
        ("Survivorship Priority 2", "Highest record_quality_score"),
        ("Survivorship Priority 3", "Highest completeness_score"),
        ("Survivorship Priority 4", "Fewest DQ issues"),
        ("Survivorship Priority 5", "Most recent record"),
        ("Survivorship Priority 6", "Lowest ID"),
    ], columns=["Item", "Value"])
    write_df(ws, run_info)

    wb.save(path)
    log.info("  Golden record report -> %s", path)
    return path


def run_golden_record_engine(
    dedup_data: dict[str, pd.DataFrame] | None = None,
    dq_data: dict[str, Any] | None = None,
    std_data: dict[str, pd.DataFrame] | None = None,
) -> dict[str, Any]:
    log.info("=" * 70)
    log.info("MODULE 6 - GOLDEN RECORD ENGINE")
    log.info("=" * 70)

    dedup_data = dedup_data or {}
    dq_data = dq_data or {}
    std_data = std_data or {}

    accounts = _frame_from_sources(dedup_data, None, "dedup_accounts", "dedup_accounts.csv")
    contacts = _frame_from_sources(dedup_data, None, "dedup_contacts", "dedup_contacts.csv")
    leads = _frame_from_sources(dedup_data, None, "dedup_leads", "dedup_leads.csv")

    dedup_map_accounts = _frame_from_sources(dedup_data, None, "dedup_map_accounts", "dedup_map_accounts.csv")
    dedup_map_contacts = _frame_from_sources(dedup_data, None, "dedup_map_contacts", "dedup_map_contacts.csv")
    dedup_map_leads = _frame_from_sources(dedup_data, None, "dedup_map_leads", "dedup_map_leads.csv")

    quality_accounts = _frame_from_sources(dq_data, None, "quality_scores_accounts", "quality_scores_accounts.csv")
    quality_contacts = _frame_from_sources(dq_data, None, "quality_scores_contacts", "quality_scores_contacts.csv")
    quality_leads = _frame_from_sources(dq_data, None, "quality_scores_leads", "quality_scores_leads.csv")

    std_golden_map = _frame_from_sources(std_data, None, "std_golden_map", "std_golden_map.csv")

    log.info("  Building golden accounts")
    golden_accounts = _build_golden_records_for_entity(
        "accounts", accounts, dedup_map_accounts, quality_accounts, std_golden_map,
    )
    log.info("    Golden accounts: %s", f"{len(golden_accounts):,}")

    log.info("  Building golden contacts")
    golden_contacts = _build_golden_records_for_entity(
        "contacts", contacts, dedup_map_contacts, quality_contacts, std_golden_map,
    )
    log.info("    Golden contacts: %s", f"{len(golden_contacts):,}")

    log.info("  Building golden leads")
    golden_leads = _build_golden_records_for_entity(
        "leads", leads, dedup_map_leads, quality_leads, std_golden_map,
    )
    log.info("    Golden leads: %s", f"{len(golden_leads):,}")

    accounts_path = _csv_path("golden_accounts.csv")
    contacts_path = _csv_path("golden_contacts.csv")
    leads_path = _csv_path("golden_leads.csv")

    golden_accounts.to_csv(accounts_path, index=False)
    golden_contacts.to_csv(contacts_path, index=False)
    golden_leads.to_csv(leads_path, index=False)

    log.info("  golden_accounts.csv -> %s", accounts_path)
    log.info("  golden_contacts.csv -> %s", contacts_path)
    log.info("  golden_leads.csv    -> %s", leads_path)

    summary_df = _summary_kpis(
        accounts, contacts, leads,
        golden_accounts, golden_contacts, golden_leads,
    )
    report_path = write_golden_record_report(
        summary_df, golden_accounts, golden_contacts, golden_leads,
    )

    total_merged = summary_df.loc[summary_df["KPI"].eq("Records Merged"), "Value"].iloc[0]
    avg_quality = summary_df.loc[summary_df["KPI"].eq("Average Quality Score"), "Value"].iloc[0]
    log.info("-" * 70)
    log.info("GOLDEN RECORD SUMMARY")
    log.info("  Accounts: %s -> %s golden", f"{len(accounts):,}", f"{len(golden_accounts):,}")
    log.info("  Contacts: %s -> %s golden", f"{len(contacts):,}", f"{len(golden_contacts):,}")
    log.info("  Leads   : %s -> %s golden", f"{len(leads):,}", f"{len(golden_leads):,}")
    log.info("  Records merged: %s", f"{total_merged:,}" if isinstance(total_merged, int) else total_merged)
    log.info("  Average quality score: %s", avg_quality)
    log.info("=" * 70)
    log.info("MODULE 6 COMPLETE")
    log.info("=" * 70)

    return {
        "golden_accounts": golden_accounts,
        "golden_contacts": golden_contacts,
        "golden_leads": golden_leads,
        "output_paths": {
            "golden_accounts": accounts_path,
            "golden_contacts": contacts_path,
            "golden_leads": leads_path,
            "golden_record_report": report_path,
            "module6_log": os.path.join(LOGS_DIR, "module6_golden_record.log"),
        },
    }


if __name__ == "__main__":
    run_golden_record_engine()
